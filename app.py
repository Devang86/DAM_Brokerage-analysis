"""
Brokerage Data Analytics
KKC & Associates LLP — Stock Broker Audit Analytics Tool
Analyzes brokerage data for anomalies, statutory compliance, and audit flags.
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
import math
from collections import Counter

from constants import (
    APP_NAME, APP_VERSION, APP_FULL_NAME,
    KKC_GREEN, KKC_GREY, WHITE, LIGHT_GREEN, LIGHT_GREEN_BG,
    LIGHT_RED, LIGHT_YELLOW, LIGHT_BLUE, LIGHT_GREY, DARK_TEXT,
    RED_ACCENT, AMBER_ACCENT, BLUE_ACCENT, FONT_NAME,
    STT_RATES, SEBI_FEE_RATE, GST_RATE,
    STAMP_DUTY_RATES, TURNOVER_CHARGE_RATES,
    DEFAULT_MATERIALITY, VARIANCE_FLAG_PCT, CLIENT_CONCENTRATION_PCT,
    ERROR_RATE_FLAG_PCT, BROKERAGE_RATE_OUTLIER_STD, ADJUSTMENT_FLAG_PCT,
    MONTH_END_BUNCHING_PCT, VOLUME_SPIKE_STD,
    COL_CLIENT_CODE, COL_CLIENT_NAME, COL_SCRIP_CODE, COL_SCRIP_NAME,
    COL_EXCHANGE, COL_TXN_DATE, COL_QTY, COL_BUY_SELL,
    COL_MARKET_VALUE, COL_NET_VALUE, COL_TOTAL_BROKERAGE, COL_STT,
    COL_CGST, COL_SGST, COL_IGST, COL_TURNOVER_TAX,
    COL_STAMP_CHARGES, COL_SEBI_FEES, COL_NET_BROKERAGE, COL_TURNOVER,
    COL_FAMILY_CODE, COL_ISIN, COL_CUSTODIAN, COL_SETTLEMENT,
    COL_BOOK_TYPE, COL_INSTRUMENT, COL_EXPIRY, COL_STRIKE, COL_OPTION_TYPE,
    COL_ALIASES, EXCEL_SHEETS, AUDIT_PROCEDURES,
    COL_PRODUCT_DESC, SEBI_MAX_BROKERAGE_PCT, CGST_RATE, SGST_RATE,
    FY2026_HOLIDAYS,
)

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Brokerage Data Analytics — KKC & Associates LLP",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# KKC Brand CSS
# ---------------------------------------------------------------------------
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Source+Sans+Pro:wght@300;400;600;700&display=swap');
    * { font-family: 'Source Sans Pro', sans-serif !important; }
    .main-header {
        background: linear-gradient(135deg, #7CB542 0%, #5a8a30 100%);
        padding: 1.5rem 2rem; border-radius: 10px; margin-bottom: 1.5rem;
        color: white;
    }
    .main-header h1 { color: white !important; margin: 0; font-size: 1.8rem; }
    .main-header p { color: #e8f5e0 !important; margin: 0.3rem 0 0 0; font-size: 1rem; }
    .metric-card {
        background: #f8f9fa; border-left: 4px solid #7CB542;
        padding: 1rem 1.2rem; border-radius: 6px; margin-bottom: 0.8rem;
    }
    .metric-card h3 { margin: 0; color: #808285; font-size: 0.85rem; text-transform: uppercase; }
    .metric-card .value { font-size: 1.6rem; font-weight: 700; color: #333; }
    .flag-high { background: #fff0f0; border-left: 4px solid #e74c3c; padding: 0.8rem; border-radius: 6px; margin: 0.4rem 0; }
    .flag-medium { background: #fff8e1; border-left: 4px solid #f39c12; padding: 0.8rem; border-radius: 6px; margin: 0.4rem 0; }
    .flag-low { background: #f0f8ff; border-left: 4px solid #3498db; padding: 0.8rem; border-radius: 6px; margin: 0.4rem 0; }
    .section-header {
        background: #f0f7e8; padding: 0.8rem 1.2rem; border-radius: 6px;
        border-left: 4px solid #7CB542; margin: 1.5rem 0 1rem 0;
    }
    .section-header h2 { margin: 0; color: #333; font-size: 1.2rem; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] {
        background: #f8f9fa; border-radius: 6px 6px 0 0;
        padding: 8px 20px; font-weight: 600;
    }
    .stTabs [aria-selected="true"] { background: #7CB542 !important; color: white !important; }
    .footer-text { text-align: center; color: #808285; font-size: 0.8rem; margin-top: 3rem; padding: 1rem; border-top: 1px solid #e0e0e0; }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------------
st.markdown("""
<div class="main-header">
    <h1>Brokerage Data Analytics</h1>
    <p>KKC & Associates LLP — Stock Broker Audit Analytics</p>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def format_inr(value):
    """Format a number in Indian numbering system (lakhs/crores)."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "-"
    negative = value < 0
    value = abs(value)
    if value >= 1e7:
        return f"{'(-)' if negative else ''}Rs. {value / 1e7:,.2f} Cr"
    elif value >= 1e5:
        return f"{'(-)' if negative else ''}Rs. {value / 1e5:,.2f} L"
    else:
        return f"{'(-)' if negative else ''}Rs. {value:,.2f}"


def format_inr_plain(value):
    """Format with commas in Indian style (no Cr/L suffix)."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "-"
    negative = value < 0
    value = abs(value)
    s = f"{value:,.2f}"
    return f"({s})" if negative else s


def parse_date_column(df, col):
    """Try multiple date formats to parse a column."""
    for fmt in [None, "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d",
                "%d-%b-%Y", "%d %b %Y", "%d %B %Y"]:
        try:
            if fmt:
                return pd.to_datetime(df[col], format=fmt, dayfirst=True)
            else:
                return pd.to_datetime(df[col], dayfirst=True)
        except Exception:
            continue
    return pd.to_datetime(df[col], errors="coerce", dayfirst=True)


def safe_div(num, den):
    """Safe division returning 0 for zero denominator."""
    try:
        if den == 0 or pd.isna(den):
            return 0.0
    except (ValueError, TypeError):
        pass
    if den == 0:
        return 0.0
    return num / den


def compute_brokerage_rate(brokerage, turnover):
    """Compute brokerage rate as brokerage / turnover."""
    return np.where(turnover > 0, brokerage / turnover, 0.0)


def compute_hhi(shares):
    """Herfindahl-Hirschman Index from percentage shares."""
    return float((shares ** 2).sum())


def classify_hhi(hhi):
    """Classify HHI score."""
    if hhi < 1500:
        return "Unconcentrated"
    elif hhi < 2500:
        return "Moderate Concentration"
    else:
        return "Highly Concentrated"


def flag_outliers_iqr(series, multiplier=1.5):
    """IQR-based outlier detection."""
    q1 = series.quantile(0.25)
    q3 = series.quantile(0.75)
    iqr = q3 - q1
    lower = q1 - multiplier * iqr
    upper = q3 + multiplier * iqr
    return (series < lower) | (series > upper)


def flag_outliers_zscore(series, threshold=2):
    """Z-score based outlier detection."""
    mean = series.mean()
    std = series.std()
    if std == 0:
        return pd.Series(False, index=series.index)
    return ((series - mean).abs() / std) > threshold


def mom_variance(series):
    """Month-on-month percentage change."""
    return series.pct_change() * 100


def metric_card(label, value):
    """Render a KKC-branded metric card."""
    return f"""
    <div class="metric-card">
        <h3>{label}</h3>
        <div class="value">{value}</div>
    </div>
    """


def section_header(title):
    """Render a section header."""
    st.markdown(f'<div class="section-header"><h2>{title}</h2></div>',
                unsafe_allow_html=True)


def flag_card(severity, text):
    """Render a flag card (high/medium/low)."""
    st.markdown(f'<div class="flag-{severity}">{text}</div>',
                unsafe_allow_html=True)


def find_column(df, target):
    """Find a column in the DataFrame using aliases."""
    if target in df.columns:
        return target
    aliases = COL_ALIASES.get(target, [])
    for alias in aliases:
        if alias in df.columns:
            return alias
    # Fuzzy match: case-insensitive substring
    target_lower = target.lower()
    for col in df.columns:
        if target_lower in str(col).lower():
            return col
    return None


def detect_sheet_type(sheet_name):
    """Classify a sheet by its name."""
    sn = sheet_name.lower()
    if "summary" in sn:
        return "summary"
    if "turnover" in sn and "error" not in sn:
        return "turnover"
    if "error" in sn or "cancel" in sn or "errror" in sn:
        return "error"
    is_fo = "fo" in sn or "f&o" in sn or "f & o" in sn
    if "ro" in sn or "trade" in sn:
        return "fo_trade" if is_fo else "cash_trade"
    # Sheets like "NSE FO Apr-Mar26" without RO/trade keyword — treat as FO trade
    if is_fo:
        return "fo_trade"
    return "other"


# ═══════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown(f"### {APP_NAME}")
    st.markdown(f"*{APP_VERSION}*")
    st.markdown("---")

    entity_name = st.text_input("Entity Name", value="DAM Capital Advisors Limited")
    audit_period = st.text_input("Audit Period", value="FY 2025-26")

    st.markdown("---")
    st.markdown("### Upload Trade Files")
    uploaded_cash = st.file_uploader(
        "1. Cash RO File",
        type=["xlsx", "xls"],
        help="Upload Excel file containing Cash segment RO (Register of Orders) data. "
             "All sheets in this file will be treated as Cash trades.",
    )
    uploaded_fo = st.file_uploader(
        "2. F&O RO File",
        type=["xlsx", "xls"],
        help="Upload Excel file containing F&O segment RO data. "
             "All sheets in this file will be treated as F&O trades.",
    )
    uploaded_error = st.file_uploader(
        "3. Error Trade File",
        type=["xlsx", "xls"],
        help="Upload Excel file containing Error / Cancelled trade data.",
    )

    st.markdown("---")
    st.markdown("### Optional Files")
    uploaded_summary = st.file_uploader(
        "Brokerage Summary (Optional)",
        type=["xlsx", "xls"],
        help="Upload the Brokerage Summary sheet for adjustment analysis and turnover reconciliation.",
    )
    uploaded_stamp = st.file_uploader(
        "Stamp Duty Bank Payment (Optional)",
        type=["xlsx", "xls"],
        help="Upload the stamp duty bank payment file for cross-check.",
    )

    st.markdown("---")
    st.markdown("### Thresholds")
    materiality = st.number_input("Materiality (Rs.)", value=DEFAULT_MATERIALITY,
                                  step=100000, format="%d")
    variance_pct = st.number_input("MoM Variance Flag (%)", value=VARIANCE_FLAG_PCT,
                                   step=5, min_value=5, max_value=100)
    concentration_pct = st.number_input("Client Concentration Flag (%)",
                                        value=CLIENT_CONCENTRATION_PCT,
                                        step=1, min_value=1, max_value=50)
    top_n = st.number_input("Top N for Reports", value=20, step=5, min_value=5, max_value=100)

    st.markdown("---")
    st.markdown(
        '<p style="color: #808285; font-size: 0.75rem;">'
        'KKC & Associates LLP<br>Chartered Accountants</p>',
        unsafe_allow_html=True,
    )


# ═══════════════════════════════════════════════════════════════════════════
# MAIN — FILE NOT UPLOADED
# ═══════════════════════════════════════════════════════════════════════════

if uploaded_cash is None and uploaded_fo is None:
    st.markdown("""
    ### How to Use
    1. Upload the **Cash RO File** (all sheets treated as Cash trades)
    2. Upload the **F&O RO File** (all sheets treated as F&O trades)
    3. Upload the **Error Trade File** (all sheets treated as error/cancelled trades)
    4. Optionally upload the **Brokerage Summary** and **Stamp Duty Bank Payment** files
    5. Adjust thresholds if needed
    6. Review results in tabs and download the Excel report

    **Supported analytics:**
    Summary Dashboard, Turnover Reconciliation, Client Concentration,
    Scrip Analysis, Brokerage Rate Analytics, Error Trade Analysis,
    GST Analytics, STT Verification, SEBI Fees, Stamp Duty Analysis,
    Adjustment Analysis, Buy/Sell Analysis, Temporal Analysis,
    Client × Product Analysis, Transaction-Level Analysis
    """)
    st.stop()


# ═══════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def load_all_sheets(file_bytes, segment_tag=""):
    """Load ALL sheets from an Excel file and concatenate.
    Each row is tagged with _source_sheet and _file_segment."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    dfs = []
    for sn in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sn, engine="openpyxl")
            if df.empty or len(df.columns) < 3:
                continue
            df["_source_sheet"] = sn
            df["_file_segment"] = segment_tag
            dfs.append(df)
        except Exception:
            pass
    xls.close()
    if dfs:
        return pd.concat(dfs, ignore_index=True)
    return pd.DataFrame()


@st.cache_data(show_spinner=False)
def load_first_sheet(file_bytes):
    """Load the first sheet from an Excel file."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], engine="openpyxl")
    xls.close()
    return df


prog = st.progress(0, text="Loading Cash RO file...")
trade_frames = []
sheet_counts = {}

# Cash RO
if uploaded_cash is not None:
    cash_bytes = uploaded_cash.read()
    df_cash = load_all_sheets(cash_bytes, segment_tag="Cash")
    if not df_cash.empty:
        trade_frames.append(df_cash)
        cash_sheets = df_cash["_source_sheet"].nunique()
        sheet_counts["Cash"] = cash_sheets
        st.success(f"Cash RO: {len(df_cash):,} rows across {cash_sheets} sheet(s) "
                   f"from {uploaded_cash.name}")

prog.progress(30, text="Loading F&O RO file...")

# F&O RO
if uploaded_fo is not None:
    fo_bytes = uploaded_fo.read()
    df_fo = load_all_sheets(fo_bytes, segment_tag="F&O")
    if not df_fo.empty:
        trade_frames.append(df_fo)
        fo_sheets = df_fo["_source_sheet"].nunique()
        sheet_counts["F&O"] = fo_sheets
        st.success(f"F&O RO: {len(df_fo):,} rows across {fo_sheets} sheet(s) "
                   f"from {uploaded_fo.name}")

prog.progress(60, text="Loading Error Trade file...")

# Error Trades
df_errors = pd.DataFrame()
if uploaded_error is not None:
    err_bytes = uploaded_error.read()
    df_errors = load_all_sheets(err_bytes, segment_tag="Error")
    if not df_errors.empty:
        err_sheets = df_errors["_source_sheet"].nunique()
        sheet_counts["Error"] = err_sheets
        st.success(f"Error Trades: {len(df_errors):,} rows across {err_sheets} sheet(s) "
                   f"from {uploaded_error.name}")

prog.progress(80, text="Loading optional files...")

# Combine trade frames
if trade_frames:
    df_trades = pd.concat(trade_frames, ignore_index=True)
else:
    df_trades = pd.DataFrame()

# Summary (optional)
df_summary_raw = pd.DataFrame()
df_turnover_raw = pd.DataFrame()
if uploaded_summary is not None:
    summary_bytes = uploaded_summary.read()
    xls_sum = pd.ExcelFile(io.BytesIO(summary_bytes), engine="openpyxl")
    # Auto-detect summary and turnover sheets
    for sn in xls_sum.sheet_names:
        sn_lower = sn.lower()
        if "summary" in sn_lower and df_summary_raw.empty:
            df_summary_raw = pd.read_excel(xls_sum, sheet_name=sn, engine="openpyxl")
        elif "turnover" in sn_lower and "error" not in sn_lower and df_turnover_raw.empty:
            df_turnover_raw = pd.read_excel(xls_sum, sheet_name=sn, engine="openpyxl")
    # If no auto-detect, use first sheet as summary
    if df_summary_raw.empty:
        df_summary_raw = pd.read_excel(xls_sum, sheet_name=xls_sum.sheet_names[0], engine="openpyxl")
    xls_sum.close()
    st.success(f"Summary file loaded from {uploaded_summary.name}")

# Stamp duty bank payment (optional)
df_stamp_bank = None
if uploaded_stamp is not None:
    stamp_bytes = uploaded_stamp.read()
    df_stamp_bank = pd.read_excel(io.BytesIO(stamp_bytes), engine="openpyxl")

prog.progress(100, text="All data loaded!")


# ---------------------------------------------------------------------------
# Prepare trade data
# ---------------------------------------------------------------------------

def prepare_trades(df):
    """Clean and enrich trade-level data."""
    if df.empty:
        return df

    # Find key columns
    col_map = {}
    for target in [COL_CLIENT_CODE, COL_CLIENT_NAME, COL_SCRIP_CODE, COL_SCRIP_NAME,
                   COL_EXCHANGE, COL_TXN_DATE, COL_QTY, COL_BUY_SELL,
                   COL_MARKET_VALUE, COL_TOTAL_BROKERAGE, COL_STT,
                   COL_CGST, COL_SGST, COL_IGST, COL_TURNOVER_TAX,
                   COL_STAMP_CHARGES, COL_SEBI_FEES, COL_NET_BROKERAGE,
                   COL_TURNOVER, COL_FAMILY_CODE, COL_ISIN, COL_SETTLEMENT,
                   COL_INSTRUMENT, COL_EXPIRY, COL_STRIKE, COL_OPTION_TYPE,
                   COL_PRODUCT_DESC]:
        found = find_column(df, target)
        if found:
            col_map[target] = found

    # Rename to standard names
    reverse_map = {v: k for k, v in col_map.items()}
    df = df.rename(columns=reverse_map)

    # Parse dates
    if COL_TXN_DATE in df.columns:
        df[COL_TXN_DATE] = parse_date_column(df, COL_TXN_DATE)
        df["Month"] = df[COL_TXN_DATE].dt.to_period("M")
        df["Month_str"] = df[COL_TXN_DATE].dt.strftime("%b'%y")
        df["Day_of_week"] = df[COL_TXN_DATE].dt.day_name()
        df["Day_of_month"] = df[COL_TXN_DATE].dt.day
        df["Days_in_month"] = df[COL_TXN_DATE].dt.days_in_month

    # Ensure numeric columns
    for col in [COL_TOTAL_BROKERAGE, COL_STT, COL_CGST, COL_SGST, COL_IGST,
                COL_TURNOVER_TAX, COL_STAMP_CHARGES, COL_SEBI_FEES,
                COL_NET_BROKERAGE, COL_TURNOVER, COL_MARKET_VALUE, COL_QTY]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # GST Total
    gst_cols = [c for c in [COL_CGST, COL_SGST, COL_IGST] if c in df.columns]
    if gst_cols:
        df["GST_Total"] = df[gst_cols].sum(axis=1)
    else:
        df["GST_Total"] = 0

    # Segment classification (vectorized for speed on large datasets)
    df["Segment"] = classify_segment_vectorized(df)

    # Brokerage rate
    if COL_TOTAL_BROKERAGE in df.columns and COL_TURNOVER in df.columns:
        df["Brokerage_Rate"] = compute_brokerage_rate(
            df[COL_TOTAL_BROKERAGE], df[COL_TURNOVER]
        )
    else:
        df["Brokerage_Rate"] = 0

    # Normalize Product Description
    if COL_PRODUCT_DESC in df.columns:
        df[COL_PRODUCT_DESC] = df[COL_PRODUCT_DESC].astype(str).str.strip().str.upper()

    return df


def classify_segment_vectorized(df):
    """Classify trades into NSE Cash, BSE Cash, or F&O using vectorized ops.
    If _file_segment is 'F&O', all trades in that file are forced to F&O.
    If _file_segment is 'Cash', classification is by exchange (NSE/BSE)."""
    segment = pd.Series("NSE Cash", index=df.index)

    # 0. If _file_segment tag exists, use it as the primary classifier
    if "_file_segment" in df.columns:
        is_fo_file = df["_file_segment"].astype(str).str.upper() == "F&O"
        segment = segment.where(~is_fo_file, "F&O")

    # 1. Source sheet name contains 'fo' -> F&O (fallback for mixed files)
    if "_source_sheet" in df.columns:
        source_lower = df["_source_sheet"].astype(str).str.lower()
        fo_source = source_lower.str.contains("fo|f&o", na=False, regex=True)
        segment = segment.where(~fo_source, "F&O")

    # 2. Instrument column contains FUT or OPT -> F&O
    if COL_INSTRUMENT in df.columns:
        instr_upper = df[COL_INSTRUMENT].astype(str).str.upper()
        fo_instr = instr_upper.str.contains("FUT|OPT", na=False, regex=True)
        segment = segment.where(~fo_instr, "F&O")

    # 3. Scrip name contains FUT or OPT -> F&O
    if COL_SCRIP_NAME in df.columns:
        scrip_upper = df[COL_SCRIP_NAME].astype(str).str.upper()
        fo_scrip = scrip_upper.str.contains("FUT|OPT", na=False, regex=True)
        segment = segment.where(~fo_scrip, "F&O")

    # 4. Expiry date is not empty -> F&O
    if COL_EXPIRY in df.columns:
        expiry_str = df[COL_EXPIRY].astype(str).str.strip()
        has_expiry = ~expiry_str.isin(["", "nan", "None", "NaT", "NaN", "nat"])
        segment = segment.where(~has_expiry, "F&O")

    # 5. Exchange is BSE and not already F&O -> BSE Cash
    if COL_EXCHANGE in df.columns:
        is_bse = df[COL_EXCHANGE].astype(str).str.upper() == "BSE"
        is_not_fo = segment != "F&O"
        segment = segment.where(~(is_bse & is_not_fo), "BSE Cash")

    return segment


df_trades = prepare_trades(df_trades)
df_errors = prepare_trades(df_errors)

total_trades = len(df_trades)
total_errors = len(df_errors)
error_rate = safe_div(total_errors, total_trades) * 100

trade_sheet_total = sum(sheet_counts.get(k, 0) for k in ["Cash", "F&O"])
error_sheet_total = sheet_counts.get("Error", 0)
st.info(f"Loaded **{total_trades:,}** trades across **{trade_sheet_total}** sheet(s) "
        f"(Cash: {sheet_counts.get('Cash', 0)}, F&O: {sheet_counts.get('F&O', 0)}) "
        f"| **{total_errors:,}** error trades across **{error_sheet_total}** sheet(s)")


# ═══════════════════════════════════════════════════════════════════════════
# PARSE SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════════════════

def parse_summary_sheet(df_raw):
    """Extract monthly brokerage data from the Summary sheet."""
    records = []
    current_segment = None

    for idx, row in df_raw.iterrows():
        first_val = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""

        # Detect segment headers
        if "NSE Cash" in first_val and "segment" in first_val.lower():
            current_segment = "NSE Cash"
            continue
        elif "BSE Cash" in first_val and "segment" in first_val.lower():
            current_segment = "BSE Cash"
            continue
        elif "F & O" in first_val or "F&O" in first_val or "FO" in first_val.upper():
            if "segment" in first_val.lower():
                current_segment = "F&O"
                continue
        elif "Consolidated" in first_val:
            current_segment = "Consolidated Cash"
            continue
        elif "Total" in first_val:
            continue

        if current_segment is None:
            continue

        # Check if this is a month row (has a date or month name)
        cell0 = row.iloc[0]
        month_label = None

        if isinstance(cell0, datetime):
            month_label = cell0.strftime("%b'%y")
        elif isinstance(cell0, str):
            cell0_clean = cell0.strip()
            if cell0_clean.startswith("Adj"):
                # Adjustment row
                gross = pd.to_numeric(row.iloc[1], errors="coerce") if len(row) > 1 else 0
                net = pd.to_numeric(row.iloc[14], errors="coerce") if len(row) > 14 else 0
                if pd.notna(gross) and gross != 0:
                    records.append({
                        "Segment": current_segment,
                        "Month": f"Adj-{len(records)}",
                        "Is_Adj": True,
                        "Gross_Brokerage": gross if pd.notna(gross) else 0,
                        "Net_Brokerage": net if pd.notna(net) else 0,
                    })
                continue
            # Check for month abbreviations like Oct'25, Nov'25
            month_names = ["jan", "feb", "mar", "apr", "may", "jun",
                           "jul", "aug", "sep", "oct", "nov", "dec"]
            for mn in month_names:
                if cell0_clean.lower().startswith(mn):
                    month_label = cell0_clean
                    break

        if month_label is None:
            continue

        # Extract values (assuming standard column order from the Summary sheet)
        # B=Gross, C=Turnover Chg, D=Recovered TC, E=Net TC,
        # F=Stamp Duty, G=Recovered SD, H=Net SD,
        # I=GST, J=Recovered GST, K=Net GST,
        # L=STT, M=SEBI, N=Total Statutory, O=Net Brokerage
        def get_val(col_idx):
            if col_idx < len(row):
                return pd.to_numeric(row.iloc[col_idx], errors="coerce")
            return 0

        records.append({
            "Segment": current_segment,
            "Month": month_label,
            "Is_Adj": False,
            "Gross_Brokerage": get_val(1) or 0,
            "Turnover_Charges_Gross": get_val(2) or 0,
            "Turnover_Charges_Recovered": get_val(3) or 0,
            "Net_Turnover_Charges": get_val(4) or 0,
            "Stamp_Duty_Gross": get_val(5) or 0,
            "Stamp_Duty_Recovered": get_val(6) or 0,
            "Net_Stamp_Duty": get_val(7) or 0,
            "GST_Gross": get_val(8) or 0,
            "GST_Recovered": get_val(9) or 0,
            "Net_GST": get_val(10) or 0,
            "STT": get_val(11) or 0,
            "SEBI_Fees": get_val(12) or 0,
            "Total_Statutory": get_val(13) or 0,
            "Net_Brokerage": get_val(14) or 0,
        })

    return pd.DataFrame(records)


df_summary = parse_summary_sheet(df_summary_raw)


# ═══════════════════════════════════════════════════════════════════════════
# STORE ALL FLAGS FOR CONSOLIDATED VIEW
# ═══════════════════════════════════════════════════════════════════════════
all_flags = []   # list of (severity, section, message)


# ═══════════════════════════════════════════════════════════════════════════
# TABS
# ═══════════════════════════════════════════════════════════════════════════

tabs = st.tabs([
    "Summary",            # 0
    "Turnover Recon",     # 1
    "Client Conc.",       # 2
    "Scrip Analysis",     # 3
    "Brokerage Rates",    # 4
    "Error Trades",       # 5
    "GST Analytics",      # 6
    "STT Verification",   # 7
    "SEBI Fees",          # 8
    "Stamp Duty",         # 9
    "Adjustments",        # 10
    "Buy/Sell",           # 11
    "Temporal",           # 12
    "Client × Product",   # 13
    "Txn Analysis",       # 14
    "Audit Flags",        # 15
])


# ═══════════════════════════════════════════════════════════════════════════
# TAB 0: SUMMARY DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════

with tabs[0]:
    section_header("Section 1: Summary Dashboard")

    # Compute from trade data
    total_gross = df_trades[COL_TOTAL_BROKERAGE].sum() if COL_TOTAL_BROKERAGE in df_trades.columns else 0
    total_net = df_trades[COL_NET_BROKERAGE].sum() if COL_NET_BROKERAGE in df_trades.columns else 0
    total_turnover = df_trades[COL_TURNOVER].sum() if COL_TURNOVER in df_trades.columns else 0
    avg_brokerage = safe_div(total_gross, total_trades)
    net_margin = safe_div(total_net, total_gross) * 100

    # Metric cards row 1
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(metric_card("Total Gross Brokerage", format_inr(total_gross)),
                    unsafe_allow_html=True)
    with c2:
        st.markdown(metric_card("Total Net Brokerage", format_inr(total_net)),
                    unsafe_allow_html=True)
    with c3:
        st.markdown(metric_card("Total Trades", f"{total_trades:,}"),
                    unsafe_allow_html=True)
    with c4:
        st.markdown(metric_card("Avg. Brokerage/Trade", format_inr(avg_brokerage)),
                    unsafe_allow_html=True)

    # Metric cards row 2
    c5, c6, c7, c8 = st.columns(4)
    with c5:
        st.markdown(metric_card("Total Turnover", format_inr(total_turnover)),
                    unsafe_allow_html=True)
    with c6:
        st.markdown(metric_card("Net Brokerage Margin", f"{net_margin:.1f}%"),
                    unsafe_allow_html=True)
    with c7:
        total_stt = df_trades[COL_STT].sum() if COL_STT in df_trades.columns else 0
        st.markdown(metric_card("Total STT", format_inr(total_stt)),
                    unsafe_allow_html=True)
    with c8:
        total_gst = df_trades["GST_Total"].sum() if "GST_Total" in df_trades.columns else 0
        st.markdown(metric_card("Total GST", format_inr(total_gst)),
                    unsafe_allow_html=True)

    # Monthly trend
    if "Month" in df_trades.columns and not df_trades.empty:
        agg_dict = {
            "Gross_Brokerage": (COL_TOTAL_BROKERAGE, "sum"),
            "Trade_Count": (COL_TOTAL_BROKERAGE, "count"),
        }
        agg_dict["Net_Brokerage"] = (COL_NET_BROKERAGE, "sum") if COL_NET_BROKERAGE in df_trades.columns else (COL_TOTAL_BROKERAGE, "sum")
        agg_dict["Turnover"] = (COL_TURNOVER, "sum") if COL_TURNOVER in df_trades.columns else (COL_TOTAL_BROKERAGE, "count")
        monthly = df_trades.groupby(["Month", "Segment"]).agg(**agg_dict).reset_index()

        monthly_pivot = monthly.pivot_table(
            index="Month", columns="Segment",
            values="Gross_Brokerage", aggfunc="sum", fill_value=0
        )
        monthly_pivot["Total"] = monthly_pivot.sum(axis=1)
        monthly_pivot = monthly_pivot.sort_index()

        st.markdown("#### Monthly Brokerage Trend")
        chart_df = monthly_pivot.copy()
        chart_df.index = chart_df.index.astype(str)
        st.line_chart(chart_df)

        # MoM Variance
        if "Total" in monthly_pivot.columns and len(monthly_pivot) > 1:
            var_series = mom_variance(monthly_pivot["Total"])
            flagged_months = var_series[var_series.abs() > variance_pct]
            if not flagged_months.empty:
                st.markdown("#### MoM Variance Flags")
                for period, pct in flagged_months.items():
                    direction = "increase" if pct > 0 else "decrease"
                    msg = f"**{period}**: {abs(pct):.1f}% {direction} from previous month"
                    flag_card("medium", msg)
                    all_flags.append(("medium", "Summary Dashboard",
                                     f"{period}: {abs(pct):.1f}% MoM {direction}"))

        # Segment contribution
        st.markdown("#### Segment Contribution")
        seg_summary = df_trades.groupby("Segment")[COL_TOTAL_BROKERAGE].sum().reset_index()
        seg_summary.columns = ["Segment", "Gross Brokerage"]
        seg_summary["% of Total"] = (seg_summary["Gross Brokerage"] / seg_summary["Gross Brokerage"].sum() * 100).round(2)
        seg_summary["Gross Brokerage (Formatted)"] = seg_summary["Gross Brokerage"].apply(format_inr)
        st.dataframe(seg_summary[["Segment", "Gross Brokerage (Formatted)", "% of Total"]],
                     hide_index=True, use_container_width=True)

    # Audit procedures
    with st.expander("Audit Procedures — Summary Dashboard"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["Summary Dashboard"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 1: TURNOVER RECONCILIATION
# ═══════════════════════════════════════════════════════════════════════════

with tabs[1]:
    section_header("Section 2: Turnover Reconciliation")

    if not df_turnover_raw.empty:
        st.markdown("#### Raw Turnover Data")
        st.dataframe(df_turnover_raw, use_container_width=True, height=400)

        # Try to parse structured turnover data
        st.markdown("#### Turnover Reconciliation from Trade Data")
        if "Month" in df_trades.columns and COL_TURNOVER in df_trades.columns:
            trade_turnover = df_trades.groupby(["Month", "Segment"])[COL_TURNOVER].sum().reset_index()
            trade_turnover.columns = ["Month", "Segment", "Turnover (As per LD)"]
            trade_turnover["Month"] = trade_turnover["Month"].astype(str)
            trade_turnover["Turnover (Formatted)"] = trade_turnover["Turnover (As per LD)"].apply(format_inr)
            st.dataframe(trade_turnover[["Month", "Segment", "Turnover (Formatted)"]],
                         hide_index=True, use_container_width=True)

            # Flag large differences if exchange data available
            total_ld = trade_turnover["Turnover (As per LD)"].sum()
            st.markdown(f"**Total Turnover (As per LD):** {format_inr(total_ld)}")
    else:
        st.warning("No turnover sheet found. Please verify sheet mapping in sidebar.")

    with st.expander("Audit Procedures — Turnover Reconciliation"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["Turnover Recon"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 2: CLIENT CONCENTRATION
# ═══════════════════════════════════════════════════════════════════════════

with tabs[2]:
    section_header("Section 3: Client Concentration Analysis")

    if COL_CLIENT_CODE in df_trades.columns and COL_TOTAL_BROKERAGE in df_trades.columns:
        name_col = COL_CLIENT_NAME if COL_CLIENT_NAME in df_trades.columns else COL_CLIENT_CODE
        fam_col = COL_FAMILY_CODE if COL_FAMILY_CODE in df_trades.columns else None

        cl_agg = {
            "Total_Brokerage": (COL_TOTAL_BROKERAGE, "sum"),
            "Trade_Count": (COL_TOTAL_BROKERAGE, "count"),
        }
        cl_agg["Total_Turnover"] = (COL_TURNOVER, "sum") if COL_TURNOVER in df_trades.columns else (COL_TOTAL_BROKERAGE, "count")
        client_summary = df_trades.groupby([COL_CLIENT_CODE, name_col]).agg(**cl_agg).reset_index()
        client_summary = client_summary.sort_values("Total_Brokerage", ascending=False)
        client_summary["Pct_of_Total"] = (client_summary["Total_Brokerage"] /
                                           client_summary["Total_Brokerage"].sum() * 100).round(3)
        client_summary["Cumulative_Pct"] = client_summary["Pct_of_Total"].cumsum().round(3)

        total_clients = len(client_summary)
        shares_pct = client_summary["Pct_of_Total"]
        hhi = compute_hhi(shares_pct)
        hhi_class = classify_hhi(hhi)
        conc_clients = len(client_summary[client_summary["Pct_of_Total"] > concentration_pct])

        # Metrics
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(metric_card("Total Clients", f"{total_clients:,}"),
                        unsafe_allow_html=True)
        with c2:
            st.markdown(metric_card("HHI Score", f"{hhi:,.0f}"),
                        unsafe_allow_html=True)
        with c3:
            st.markdown(metric_card("HHI Classification", hhi_class),
                        unsafe_allow_html=True)
        with c4:
            st.markdown(metric_card(f"Clients > {concentration_pct}%", f"{conc_clients}"),
                        unsafe_allow_html=True)

        if hhi >= 2500:
            flag_card("high", f"Highly concentrated client base (HHI: {hhi:,.0f}). "
                      "Revenue dependency risk — document in going concern assessment.")
            all_flags.append(("high", "Client Concentration",
                             f"HHI {hhi:,.0f} — Highly concentrated"))

        if conc_clients > 0:
            flag_card("high", f"{conc_clients} client(s) contribute more than "
                      f"{concentration_pct}% each of total brokerage.")
            all_flags.append(("high", "Client Concentration",
                             f"{conc_clients} clients above {concentration_pct}% threshold"))

        # Top N clients table
        st.markdown(f"#### Top {top_n} Clients by Brokerage")
        display_df = client_summary.head(top_n).copy()
        display_df["Brokerage (Formatted)"] = display_df["Total_Brokerage"].apply(format_inr)
        display_df["% of Total"] = display_df["Pct_of_Total"].apply(lambda x: f"{x:.2f}%")
        display_df["Cumulative %"] = display_df["Cumulative_Pct"].apply(lambda x: f"{x:.2f}%")
        st.dataframe(
            display_df[[COL_CLIENT_CODE, name_col, "Trade_Count",
                        "Brokerage (Formatted)", "% of Total", "Cumulative %"]],
            hide_index=True, use_container_width=True
        )

        # Bar chart
        st.markdown("#### Top 10 Clients")
        chart_data = client_summary.head(10).set_index(name_col)["Total_Brokerage"]
        st.bar_chart(chart_data)

        # Top N by Turnover
        st.markdown(f"#### Top {top_n} Clients by Turnover")
        client_by_turnover = client_summary.sort_values("Total_Turnover", ascending=False).head(top_n).copy()
        client_by_turnover["Turnover (Formatted)"] = client_by_turnover["Total_Turnover"].apply(format_inr)
        client_by_turnover["Brokerage (Formatted)"] = client_by_turnover["Total_Brokerage"].apply(format_inr)
        client_by_turnover["Avg Rate (bps)"] = np.where(
            client_by_turnover["Total_Turnover"] > 0,
            client_by_turnover["Total_Brokerage"] / client_by_turnover["Total_Turnover"] * 10000, 0
        ).round(2)
        st.dataframe(
            client_by_turnover[[COL_CLIENT_CODE, name_col, "Trade_Count",
                                "Turnover (Formatted)", "Brokerage (Formatted)", "Avg Rate (bps)"]],
            hide_index=True, use_container_width=True
        )

        # Top N by Average Brokerage Rate
        st.markdown(f"#### Top {top_n} Clients by Avg Brokerage Rate")
        client_summary["Avg_Rate_bps"] = np.where(
            client_summary["Total_Turnover"] > 0,
            client_summary["Total_Brokerage"] / client_summary["Total_Turnover"] * 10000, 0
        ).round(2)
        meaningful = client_summary[client_summary["Trade_Count"] >= 10]
        client_by_rate = meaningful.sort_values("Avg_Rate_bps", ascending=False).head(top_n).copy()
        client_by_rate["Brokerage (Formatted)"] = client_by_rate["Total_Brokerage"].apply(format_inr)
        st.dataframe(
            client_by_rate[[COL_CLIENT_CODE, name_col, "Trade_Count",
                            "Brokerage (Formatted)", "Avg_Rate_bps"]],
            hide_index=True, use_container_width=True
        )

        # Family code aggregation
        if fam_col and fam_col in df_trades.columns:
            st.markdown("#### Family Code Aggregation")
            fam_summary = df_trades.groupby(fam_col).agg(
                Total_Brokerage=(COL_TOTAL_BROKERAGE, "sum"),
                Client_Count=(COL_CLIENT_CODE, "nunique"),
            ).reset_index().sort_values("Total_Brokerage", ascending=False)
            fam_summary["Pct"] = (fam_summary["Total_Brokerage"] /
                                  fam_summary["Total_Brokerage"].sum() * 100).round(2)
            fam_summary["Brokerage (Formatted)"] = fam_summary["Total_Brokerage"].apply(format_inr)
            st.dataframe(
                fam_summary.head(top_n)[[fam_col, "Client_Count", "Brokerage (Formatted)", "Pct"]],
                hide_index=True, use_container_width=True
            )
        # MoM Brokerage vs Turnover
        if "Month" in df_trades.columns and COL_TURNOVER in df_trades.columns:
            st.markdown("#### Month-on-Month: Brokerage vs Turnover")
            mom_agg = df_trades.groupby("Month").agg(
                Gross_Brokerage=(COL_TOTAL_BROKERAGE, "sum"),
                Turnover=(COL_TURNOVER, "sum"),
            ).reset_index().sort_index()
            mom_agg["Brokerage_MoM_%"] = mom_variance(mom_agg["Gross_Brokerage"]).round(1)
            mom_agg["Turnover_MoM_%"] = mom_variance(mom_agg["Turnover"]).round(1)
            mom_agg["Avg_Rate_bps"] = np.where(
                mom_agg["Turnover"] > 0,
                mom_agg["Gross_Brokerage"] / mom_agg["Turnover"] * 10000, 0
            ).round(2)
            mom_agg["Month"] = mom_agg["Month"].astype(str)
            for col_m in ["Gross_Brokerage", "Turnover"]:
                mom_agg[f"{col_m}_fmt"] = mom_agg[col_m].apply(format_inr)
            st.dataframe(
                mom_agg[["Month", "Gross_Brokerage_fmt", "Brokerage_MoM_%",
                         "Turnover_fmt", "Turnover_MoM_%", "Avg_Rate_bps"]],
                hide_index=True, use_container_width=True
            )
            # Flag divergence
            divergent = mom_agg[
                ((mom_agg["Brokerage_MoM_%"] > 0) & (mom_agg["Turnover_MoM_%"] < -10)) |
                ((mom_agg["Brokerage_MoM_%"] < -10) & (mom_agg["Turnover_MoM_%"] > 0))
            ].dropna(subset=["Brokerage_MoM_%", "Turnover_MoM_%"])
            if not divergent.empty:
                flag_card("medium", f"{len(divergent)} month(s) where brokerage and turnover "
                          "trends diverge — investigate rate changes or client mix shifts.")
                all_flags.append(("medium", "Client Concentration",
                                 f"{len(divergent)} months with brokerage-turnover divergence"))
    else:
        st.warning("Client code or brokerage columns not found in trade data.")

    with st.expander("Audit Procedures — Client Concentration"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["Client Concentration"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 3: SCRIP ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

with tabs[3]:
    section_header("Section 4: Scrip-wise Analysis")

    scrip_col = COL_SCRIP_NAME if COL_SCRIP_NAME in df_trades.columns else COL_SCRIP_CODE
    if scrip_col in df_trades.columns and COL_TOTAL_BROKERAGE in df_trades.columns:
        scrip_agg = {
            "Total_Brokerage": (COL_TOTAL_BROKERAGE, "sum"),
            "Trade_Count": (COL_TOTAL_BROKERAGE, "count"),
        }
        scrip_agg["Total_Turnover"] = (COL_TURNOVER, "sum") if COL_TURNOVER in df_trades.columns else (COL_TOTAL_BROKERAGE, "count")
        scrip_summary = df_trades.groupby(scrip_col).agg(**scrip_agg).reset_index()
        scrip_summary["Avg_Rate"] = compute_brokerage_rate(
            scrip_summary["Total_Brokerage"], scrip_summary["Total_Turnover"]
        )
        scrip_summary = scrip_summary.sort_values("Total_Brokerage", ascending=False)
        scrip_summary["Pct_of_Total"] = (scrip_summary["Total_Brokerage"] /
                                          scrip_summary["Total_Brokerage"].sum() * 100).round(3)

        st.markdown(f"#### Top {top_n} Scrips by Brokerage")
        display_scrip = scrip_summary.head(top_n).copy()
        display_scrip["Brokerage"] = display_scrip["Total_Brokerage"].apply(format_inr)
        display_scrip["Turnover"] = display_scrip["Total_Turnover"].apply(format_inr)
        display_scrip["Avg Rate (bps)"] = (display_scrip["Avg_Rate"] * 10000).round(2)
        st.dataframe(
            display_scrip[[scrip_col, "Trade_Count", "Turnover", "Brokerage",
                           "Avg Rate (bps)", "Pct_of_Total"]],
            hide_index=True, use_container_width=True
        )

        # Outlier rates
        st.markdown("#### Scrips with Unusual Brokerage Rates")
        valid_rates = scrip_summary[scrip_summary["Avg_Rate"] > 0].copy()
        if not valid_rates.empty:
            outlier_mask = flag_outliers_iqr(valid_rates["Avg_Rate"])
            outliers = valid_rates[outlier_mask].sort_values("Avg_Rate", ascending=False)
            if not outliers.empty:
                outliers_disp = outliers.copy()
                outliers_disp["Rate (bps)"] = (outliers_disp["Avg_Rate"] * 10000).round(2)
                outliers_disp["Brokerage"] = outliers_disp["Total_Brokerage"].apply(format_inr)
                st.dataframe(
                    outliers_disp[[scrip_col, "Trade_Count", "Brokerage", "Rate (bps)"]].head(30),
                    hide_index=True, use_container_width=True
                )
                all_flags.append(("medium", "Scrip Analysis",
                                  f"{len(outliers)} scrips with unusual brokerage rates"))
            else:
                st.info("No outlier brokerage rates detected.")
    else:
        st.warning("Scrip or brokerage columns not found.")

    with st.expander("Audit Procedures — Scrip Analysis"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["Scrip Analysis"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 4: BROKERAGE RATES
# ═══════════════════════════════════════════════════════════════════════════

with tabs[4]:
    section_header("Section 5: Brokerage Rate Analytics")

    if "Brokerage_Rate" in df_trades.columns and "Segment" in df_trades.columns:
        valid_trades = df_trades[df_trades["Brokerage_Rate"] > 0].copy()

        # Segment-wise stats
        rate_stats = valid_trades.groupby("Segment")["Brokerage_Rate"].agg(
            ["mean", "median", "min", "max", "std", "count"]
        ).reset_index()
        rate_stats.columns = ["Segment", "Mean", "Median", "Min", "Max", "Std Dev", "Count"]
        for col in ["Mean", "Median", "Min", "Max", "Std Dev"]:
            rate_stats[f"{col} (bps)"] = (rate_stats[col] * 10000).round(4)

        c1, c2, c3 = st.columns(3)
        segments = rate_stats["Segment"].tolist()
        cols = [c1, c2, c3]
        for i, seg in enumerate(segments[:3]):
            with cols[i]:
                mean_bps = rate_stats[rate_stats["Segment"] == seg]["Mean (bps)"].values
                val = f"{mean_bps[0]:.2f} bps" if len(mean_bps) > 0 else "—"
                st.markdown(metric_card(f"Avg Rate — {seg}", val),
                            unsafe_allow_html=True)

        st.markdown("#### Rate Statistics by Segment")
        st.dataframe(
            rate_stats[["Segment", "Count", "Mean (bps)", "Median (bps)",
                        "Min (bps)", "Max (bps)", "Std Dev (bps)"]],
            hide_index=True, use_container_width=True
        )

        # Unusual rate trades
        st.markdown("#### Unusual Rate Trades (Outliers)")
        outlier_trades = pd.DataFrame()
        for seg in valid_trades["Segment"].unique():
            seg_data = valid_trades[valid_trades["Segment"] == seg]
            mask = flag_outliers_zscore(seg_data["Brokerage_Rate"],
                                        threshold=BROKERAGE_RATE_OUTLIER_STD)
            outlier_trades = pd.concat([outlier_trades, seg_data[mask]])

        if not outlier_trades.empty:
            st.warning(f"{len(outlier_trades)} trades with unusual brokerage rates detected.")
            disp_cols = [c for c in [COL_CLIENT_NAME, COL_SCRIP_NAME, COL_TXN_DATE,
                                      "Segment", COL_TOTAL_BROKERAGE, COL_TURNOVER,
                                      "Brokerage_Rate"]
                         if c in outlier_trades.columns]
            outlier_disp = outlier_trades[disp_cols].head(50).copy()
            if "Brokerage_Rate" in outlier_disp.columns:
                outlier_disp["Rate (bps)"] = (outlier_disp["Brokerage_Rate"] * 10000).round(2)
            st.dataframe(outlier_disp, hide_index=True, use_container_width=True)
            all_flags.append(("medium", "Brokerage Rates",
                             f"{len(outlier_trades)} trades with unusual brokerage rates"))
        else:
            st.info("No unusual rate trades detected.")

        # Monthly rate trend
        if "Month" in valid_trades.columns:
            st.markdown("#### Monthly Average Rate Trend")
            monthly_rate = valid_trades.groupby(["Month", "Segment"])["Brokerage_Rate"].mean().reset_index()
            monthly_pivot_rate = monthly_rate.pivot_table(
                index="Month", columns="Segment", values="Brokerage_Rate", fill_value=0
            )
            monthly_pivot_rate = (monthly_pivot_rate * 10000).round(2)
            monthly_pivot_rate.index = monthly_pivot_rate.index.astype(str)
            st.line_chart(monthly_pivot_rate)
    else:
        st.warning("Brokerage rate data not available.")

    with st.expander("Audit Procedures — Brokerage Rates"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["Brokerage Rates"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 5: ERROR TRADES
# ═══════════════════════════════════════════════════════════════════════════

with tabs[5]:
    section_header("Section 6: Error/Cancelled Trade Analysis")

    error_brokerage = df_errors[COL_TOTAL_BROKERAGE].sum() if COL_TOTAL_BROKERAGE in df_errors.columns else 0

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(metric_card("Total Error Trades", f"{total_errors:,}"),
                    unsafe_allow_html=True)
    with c2:
        st.markdown(metric_card("Error Rate", f"{error_rate:.3f}%"),
                    unsafe_allow_html=True)
    with c3:
        st.markdown(metric_card("Error Brokerage", format_inr(abs(error_brokerage))),
                    unsafe_allow_html=True)
    with c4:
        cash_errors = len(df_errors[df_errors.get("Segment", pd.Series()) != "F&O"]) if "Segment" in df_errors.columns else total_errors
        fo_errors = total_errors - cash_errors
        st.markdown(metric_card("Cash / F&O Errors", f"{cash_errors} / {fo_errors}"),
                    unsafe_allow_html=True)

    if error_rate > ERROR_RATE_FLAG_PCT:
        flag_card("high", f"Error rate ({error_rate:.3f}%) exceeds {ERROR_RATE_FLAG_PCT}% threshold. "
                  "Indicates potential control weakness in order management system.")
        all_flags.append(("high", "Error Trades",
                         f"Error rate {error_rate:.3f}% exceeds threshold"))

    if not df_errors.empty:
        # Monthly error summary
        if "Month" in df_errors.columns:
            st.markdown("#### Monthly Error Summary")
            error_monthly = df_errors.groupby("Month").agg(
                Error_Count=(COL_TOTAL_BROKERAGE, "count"),
                Error_Brokerage=(COL_TOTAL_BROKERAGE, "sum"),
            ).reset_index()
            error_monthly["Month"] = error_monthly["Month"].astype(str)
            error_monthly["Brokerage (Formatted)"] = error_monthly["Error_Brokerage"].apply(
                lambda x: format_inr(abs(x))
            )
            st.dataframe(error_monthly, hide_index=True, use_container_width=True)

        # Large error trades
        if COL_TOTAL_BROKERAGE in df_errors.columns:
            st.markdown("#### Large Error Trades (Above Materiality)")
            large_errors = df_errors[df_errors[COL_TOTAL_BROKERAGE].abs() >= materiality]
            if not large_errors.empty:
                disp_cols = [c for c in [COL_CLIENT_NAME, COL_SCRIP_NAME, COL_TXN_DATE,
                                          "Segment", COL_TOTAL_BROKERAGE, COL_MARKET_VALUE]
                             if c in large_errors.columns]
                st.dataframe(large_errors[disp_cols], hide_index=True, use_container_width=True)
                all_flags.append(("high", "Error Trades",
                                 f"{len(large_errors)} error trades above materiality"))
            else:
                st.info("No error trades above materiality threshold.")

        # Pattern analysis
        if COL_CLIENT_CODE in df_errors.columns:
            st.markdown("#### Error Pattern — Repeated Clients")
            client_errors = df_errors.groupby(COL_CLIENT_CODE).size().reset_index(name="Error_Count")
            client_errors = client_errors[client_errors["Error_Count"] > 3].sort_values(
                "Error_Count", ascending=False
            )
            if not client_errors.empty:
                st.dataframe(client_errors, hide_index=True, use_container_width=True)
                all_flags.append(("medium", "Error Trades",
                                 f"{len(client_errors)} clients with repeated errors"))

        if COL_SCRIP_NAME in df_errors.columns:
            st.markdown("#### Error Pattern — Repeated Scrips")
            scrip_errors = df_errors.groupby(COL_SCRIP_NAME).size().reset_index(name="Error_Count")
            scrip_errors = scrip_errors[scrip_errors["Error_Count"] > 3].sort_values(
                "Error_Count", ascending=False
            )
            if not scrip_errors.empty:
                st.dataframe(scrip_errors, hide_index=True, use_container_width=True)
    else:
        st.info("No error trade data available.")

    with st.expander("Audit Procedures — Error Trades"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["Error Trades"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 6: GST ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════

with tabs[6]:
    section_header("Section 7: GST Analytics")

    has_gst = "GST_Total" in df_trades.columns
    if has_gst and COL_TOTAL_BROKERAGE in df_trades.columns:
        total_cgst = df_trades[COL_CGST].sum() if COL_CGST in df_trades.columns else 0
        total_sgst = df_trades[COL_SGST].sum() if COL_SGST in df_trades.columns else 0
        total_igst = df_trades[COL_IGST].sum() if COL_IGST in df_trades.columns else 0
        total_gst_all = total_cgst + total_sgst + total_igst
        interstate_pct = safe_div(total_igst, total_gst_all) * 100

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(metric_card("Total CGST", format_inr(total_cgst)),
                        unsafe_allow_html=True)
        with c2:
            st.markdown(metric_card("Total SGST", format_inr(total_sgst)),
                        unsafe_allow_html=True)
        with c3:
            st.markdown(metric_card("Total IGST", format_inr(total_igst)),
                        unsafe_allow_html=True)
        with c4:
            st.markdown(metric_card("Interstate %", f"{interstate_pct:.1f}%"),
                        unsafe_allow_html=True)

        # Rate verification — GST is 18% of Net Brokerage (taxable value)
        st.markdown("#### GST Rate Verification (Base: Net Brokerage)")
        gst_base_col = COL_NET_BROKERAGE if COL_NET_BROKERAGE in df_trades.columns else COL_TOTAL_BROKERAGE
        gst_base_label = "Net Brokerage" if gst_base_col == COL_NET_BROKERAGE else "Total Brokerage"
        if gst_base_col != COL_NET_BROKERAGE:
            st.caption(f"⚠ Net Brokerage column not found — falling back to {gst_base_label}.")
        gst_check = df_trades[df_trades[gst_base_col] > 0].copy()
        if not gst_check.empty:
            gst_check["GST_Rate"] = np.where(
                gst_check[gst_base_col] > 0,
                gst_check["GST_Total"] / gst_check[gst_base_col],
                0
            )
            tolerance = 0.005
            gst_check["Rate_OK"] = (gst_check["GST_Rate"] - GST_RATE).abs() <= tolerance
            deviations = gst_check[~gst_check["Rate_OK"]]
            if not deviations.empty:
                st.warning(f"{len(deviations)} trades with GST rate deviation "
                           f"(not 18% of {gst_base_label} +/- 0.5%)")
                disp_cols = [c for c in [COL_CLIENT_NAME, COL_SCRIP_NAME, COL_TXN_DATE,
                                          gst_base_col, "GST_Total", "GST_Rate"]
                             if c in deviations.columns]
                dev_disp = deviations[disp_cols].head(50).copy()
                if "GST_Rate" in dev_disp.columns:
                    dev_disp["GST Rate %"] = (dev_disp["GST_Rate"] * 100).round(2)
                st.dataframe(dev_disp, hide_index=True, use_container_width=True)
                all_flags.append(("medium", "GST Analytics",
                                 f"{len(deviations)} trades with GST rate deviation"))
            else:
                st.info(f"All trades have GST at 18% of {gst_base_label} (within tolerance).")

        # CGST/SGST Split Verification (intra-state trades)
        st.markdown(f"#### CGST/SGST Split Verification (9% + 9% of {gst_base_label})")
        if COL_CGST in df_trades.columns and COL_SGST in df_trades.columns and COL_IGST in df_trades.columns:
            intra = df_trades[
                (df_trades[COL_IGST] == 0) & (df_trades[gst_base_col] > 0) &
                ((df_trades[COL_CGST] > 0) | (df_trades[COL_SGST] > 0))
            ].copy()
            if not intra.empty:
                intra["CGST_Rate"] = intra[COL_CGST] / intra[gst_base_col]
                intra["SGST_Rate"] = intra[COL_SGST] / intra[gst_base_col]
                tol_gst = 0.003
                intra["CGST_OK"] = (intra["CGST_Rate"] - CGST_RATE).abs() <= tol_gst
                intra["SGST_OK"] = (intra["SGST_Rate"] - SGST_RATE).abs() <= tol_gst
                combined_dev_gst = intra[~(intra["CGST_OK"] & intra["SGST_OK"])]
                if not combined_dev_gst.empty:
                    st.warning(f"{len(combined_dev_gst)} intra-state trades with CGST/SGST rate deviation "
                               f"(expected 9%/9% of {gst_base_label})")
                    gst_split_cols = [c for c in [COL_CLIENT_NAME, COL_TXN_DATE, gst_base_col,
                                                   COL_CGST, COL_SGST, "CGST_Rate", "SGST_Rate"]
                                      if c in combined_dev_gst.columns]
                    gst_split_disp = combined_dev_gst[gst_split_cols].head(50).copy()
                    gst_split_disp["CGST %"] = (gst_split_disp["CGST_Rate"] * 100).round(2)
                    gst_split_disp["SGST %"] = (gst_split_disp["SGST_Rate"] * 100).round(2)
                    st.dataframe(gst_split_disp, hide_index=True, use_container_width=True)
                    all_flags.append(("medium", "GST Analytics",
                                     f"{len(combined_dev_gst)} trades with CGST/SGST split deviation"))
                else:
                    st.info(f"All {len(intra):,} intra-state trades have CGST = 9% and SGST = 9% (within tolerance).")
            else:
                st.info("No intra-state (CGST/SGST) trades found.")

            # IGST Verification (inter-state trades)
            st.markdown(f"#### IGST Verification (18% of {gst_base_label})")
            inter = df_trades[
                (df_trades[COL_IGST] > 0) & (df_trades[gst_base_col] > 0) &
                (df_trades[COL_CGST] == 0) & (df_trades[COL_SGST] == 0)
            ].copy()
            if not inter.empty:
                inter["IGST_Rate"] = inter[COL_IGST] / inter[gst_base_col]
                inter["IGST_OK"] = (inter["IGST_Rate"] - GST_RATE).abs() <= 0.005
                igst_dev = inter[~inter["IGST_OK"]]
                if not igst_dev.empty:
                    st.warning(f"{len(igst_dev)} inter-state trades with IGST rate deviation "
                               f"(expected 18% of {gst_base_label})")
                    igst_cols = [c for c in [COL_CLIENT_NAME, COL_TXN_DATE, gst_base_col,
                                             COL_IGST, "IGST_Rate"] if c in igst_dev.columns]
                    igst_disp = igst_dev[igst_cols].head(50).copy()
                    igst_disp["IGST %"] = (igst_disp["IGST_Rate"] * 100).round(2)
                    st.dataframe(igst_disp, hide_index=True, use_container_width=True)
                    all_flags.append(("medium", "GST Analytics",
                                     f"{len(igst_dev)} inter-state trades with IGST deviation"))
                else:
                    st.info(f"All {len(inter):,} inter-state trades have IGST at 18% (within tolerance).")
            else:
                st.info("No inter-state (IGST-only) trades found.")

        # Monthly reconciliation — GST base is Net Brokerage
        if "Month" in df_trades.columns:
            st.markdown(f"#### Monthly GST Reconciliation (Base: {gst_base_label})")
            monthly_gst = df_trades.groupby("Month").agg(
                Brokerage=(gst_base_col, "sum"),
                GST_Actual=("GST_Total", "sum"),
            ).reset_index()
            monthly_gst["GST_Expected"] = monthly_gst["Brokerage"] * GST_RATE
            monthly_gst["Difference"] = monthly_gst["GST_Actual"] - monthly_gst["GST_Expected"]
            monthly_gst["Diff_%"] = (monthly_gst["Difference"] / monthly_gst["GST_Expected"] * 100).round(2)
            monthly_gst["Month"] = monthly_gst["Month"].astype(str)
            for col in ["Brokerage", "GST_Actual", "GST_Expected", "Difference"]:
                monthly_gst[f"{col}_fmt"] = monthly_gst[col].apply(format_inr)
            st.dataframe(
                monthly_gst[["Month", "Brokerage_fmt", "GST_Expected_fmt",
                             "GST_Actual_fmt", "Difference_fmt", "Diff_%"]],
                hide_index=True, use_container_width=True
            )
    else:
        st.warning("GST data not available in trade data.")

    with st.expander("Audit Procedures — GST Analytics"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["GST Analytics"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 7: STT VERIFICATION
# ═══════════════════════════════════════════════════════════════════════════

with tabs[7]:
    section_header("Section 8: STT Verification")

    if COL_STT in df_trades.columns and COL_MARKET_VALUE in df_trades.columns:
        stt_trades = df_trades[(df_trades[COL_STT] > 0) & (df_trades[COL_MARKET_VALUE] > 0)].copy()
        stt_trades["STT_Rate"] = stt_trades[COL_STT] / stt_trades[COL_MARKET_VALUE]

        # Classify trade types
        def classify_stt_type(row):
            seg = row.get("Segment", "NSE Cash")
            side = str(row.get(COL_BUY_SELL, "")).upper()
            if "F&O" in seg:
                # Simplified — actual classification needs instrument type
                if "S" in side:
                    return "FO Sell"
                return "FO Buy"
            if "B" in side:
                return "Cash Buy"
            return "Cash Sell"

        stt_trades["STT_Type"] = stt_trades.apply(classify_stt_type, axis=1)

        st.markdown("#### STT Rate Verification by Trade Type")
        stt_summary = stt_trades.groupby("STT_Type").agg(
            Trade_Count=("STT_Rate", "count"),
            Total_STT=(COL_STT, "sum"),
            Total_Value=(COL_MARKET_VALUE, "sum"),
        ).reset_index()
        stt_summary["Effective_Rate"] = stt_summary["Total_STT"] / stt_summary["Total_Value"]
        stt_summary["Rate_bps"] = (stt_summary["Effective_Rate"] * 10000).round(4)
        stt_summary["STT (Formatted)"] = stt_summary["Total_STT"].apply(format_inr)
        st.dataframe(
            stt_summary[["STT_Type", "Trade_Count", "STT (Formatted)", "Rate_bps"]],
            hide_index=True, use_container_width=True
        )

        # Per-transaction STT rate check
        st.markdown("#### Per-Transaction STT Rate Flagging (Cash Delivery = 0.1%)")
        cash_stt = stt_trades[stt_trades["STT_Type"].isin(["Cash Buy", "Cash Sell"])].copy()
        if not cash_stt.empty:
            expected_stt_rate = STT_RATES.get("Cash Delivery Buy", 0.001)
            cash_stt["STT_Expected"] = cash_stt[COL_MARKET_VALUE] * expected_stt_rate
            cash_stt["STT_Diff"] = cash_stt[COL_STT] - cash_stt["STT_Expected"]
            stt_tol = 0.0002
            cash_stt["Rate_OK"] = (cash_stt["STT_Rate"] - expected_stt_rate).abs() <= stt_tol
            stt_devs = cash_stt[~cash_stt["Rate_OK"]]
            if not stt_devs.empty:
                st.warning(f"{len(stt_devs)} cash trades with STT rate deviation "
                           f"(expected {expected_stt_rate*100:.2f}% of Market Value)")
                stt_dev_cols = [c for c in [COL_CLIENT_NAME, COL_SCRIP_NAME, COL_TXN_DATE,
                                            COL_MARKET_VALUE, COL_STT, "STT_Rate",
                                            "STT_Expected", "STT_Diff"]
                                if c in stt_devs.columns]
                stt_dev_disp = stt_devs[stt_dev_cols].head(50).copy()
                stt_dev_disp["STT Rate %"] = (stt_dev_disp["STT_Rate"] * 100).round(4)
                st.dataframe(stt_dev_disp, hide_index=True, use_container_width=True)
                all_flags.append(("medium", "STT Verification",
                                 f"{len(stt_devs)} cash trades with STT rate deviation"))
            else:
                st.info(f"All {len(cash_stt):,} cash delivery trades have correct STT "
                        f"at {expected_stt_rate*100:.2f}%.")

        # Segment-wise STT
        st.markdown("#### Segment-wise STT Summary")
        seg_stt = df_trades.groupby("Segment")[COL_STT].sum().reset_index()
        seg_stt["STT (Formatted)"] = seg_stt[COL_STT].apply(format_inr)
        st.dataframe(seg_stt[["Segment", "STT (Formatted)"]],
                     hide_index=True, use_container_width=True)

        # Monthly trend
        if "Month" in df_trades.columns:
            st.markdown("#### Monthly STT Trend")
            monthly_stt = df_trades.groupby("Month").agg(
                STT=(COL_STT, "sum"),
                Turnover=(COL_TURNOVER, "sum") if COL_TURNOVER in df_trades.columns else (COL_MARKET_VALUE, "sum"),
            ).reset_index()
            monthly_stt["Month"] = monthly_stt["Month"].astype(str)
            st.line_chart(monthly_stt.set_index("Month")[["STT", "Turnover"]])
    else:
        st.warning("STT data not available.")

    with st.expander("Audit Procedures — STT Verification"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["STT Verification"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 8: SEBI FEES
# ═══════════════════════════════════════════════════════════════════════════

with tabs[8]:
    section_header("Section 9: SEBI Fees Verification")

    sebi_col = find_column(df_trades, COL_SEBI_FEES)
    if sebi_col and COL_TURNOVER in df_trades.columns:
        total_sebi = df_trades[sebi_col].sum()
        total_turnover_sebi = df_trades[COL_TURNOVER].sum()
        effective_rate = safe_div(total_sebi, total_turnover_sebi)
        prescribed_rate = SEBI_FEE_RATE

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(metric_card("Total SEBI Fees", format_inr(total_sebi)),
                        unsafe_allow_html=True)
        with c2:
            st.markdown(metric_card("Total Turnover", format_inr(total_turnover_sebi)),
                        unsafe_allow_html=True)
        with c3:
            st.markdown(metric_card("Effective Rate", f"{effective_rate*1e7:.2f} per Cr"),
                        unsafe_allow_html=True)
        with c4:
            st.markdown(metric_card("Prescribed Rate", f"Rs 10 per Cr"),
                        unsafe_allow_html=True)

        # Monthly verification
        if "Month" in df_trades.columns:
            st.markdown("#### Monthly SEBI Fees Verification")
            monthly_sebi = df_trades.groupby("Month").agg(
                SEBI_Fees=(sebi_col, "sum"),
                Turnover=(COL_TURNOVER, "sum"),
            ).reset_index()
            monthly_sebi["Expected_SEBI"] = monthly_sebi["Turnover"] * SEBI_FEE_RATE
            monthly_sebi["Difference"] = monthly_sebi["SEBI_Fees"] - monthly_sebi["Expected_SEBI"]
            monthly_sebi["Effective_Rate_per_Cr"] = (monthly_sebi["SEBI_Fees"] / monthly_sebi["Turnover"] * 1e7).round(2)
            monthly_sebi["Month"] = monthly_sebi["Month"].astype(str)
            for col in ["SEBI_Fees", "Expected_SEBI", "Difference"]:
                monthly_sebi[f"{col}_fmt"] = monthly_sebi[col].apply(format_inr)
            st.dataframe(
                monthly_sebi[["Month", "SEBI_Fees_fmt", "Expected_SEBI_fmt",
                              "Difference_fmt", "Effective_Rate_per_Cr"]],
                hide_index=True, use_container_width=True
            )

        # Exchange-wise
        if COL_EXCHANGE in df_trades.columns:
            st.markdown("#### Exchange-wise SEBI Fees")
            exch_sebi = df_trades.groupby(COL_EXCHANGE).agg(
                SEBI_Fees=(sebi_col, "sum"),
                Turnover=(COL_TURNOVER, "sum"),
            ).reset_index()
            exch_sebi["Rate_per_Cr"] = (exch_sebi["SEBI_Fees"] / exch_sebi["Turnover"] * 1e7).round(2)
            exch_sebi["SEBI_Fees_fmt"] = exch_sebi["SEBI_Fees"].apply(format_inr)
            st.dataframe(exch_sebi[[COL_EXCHANGE, "SEBI_Fees_fmt", "Rate_per_Cr"]],
                         hide_index=True, use_container_width=True)

        # Per-transaction SEBI fee verification
        st.markdown("#### Per-Transaction SEBI Fee Flagging (0.0001% of Market Value)")
        if COL_MARKET_VALUE in df_trades.columns:
            sebi_check = df_trades[
                (df_trades[sebi_col] > 0) & (df_trades[COL_MARKET_VALUE] > 0)
            ].copy()
            if not sebi_check.empty:
                sebi_check["SEBI_Expected"] = sebi_check[COL_MARKET_VALUE] * SEBI_FEE_RATE
                sebi_check["SEBI_Rate"] = sebi_check[sebi_col] / sebi_check[COL_MARKET_VALUE]
                sebi_check["SEBI_Diff"] = sebi_check[sebi_col] - sebi_check["SEBI_Expected"]
                sebi_tol = SEBI_FEE_RATE * 0.1
                sebi_check["Rate_OK"] = (sebi_check["SEBI_Rate"] - SEBI_FEE_RATE).abs() <= sebi_tol
                sebi_devs = sebi_check[~sebi_check["Rate_OK"]]
                if not sebi_devs.empty:
                    st.warning(f"{len(sebi_devs)} trades with SEBI fee rate deviation "
                               f"(expected 0.0001% of Market Value)")
                    sebi_dev_cols = [c for c in [COL_CLIENT_NAME, COL_SCRIP_NAME, COL_TXN_DATE,
                                                  COL_MARKET_VALUE, sebi_col,
                                                  "SEBI_Expected", "SEBI_Diff"]
                                     if c in sebi_devs.columns]
                    st.dataframe(sebi_devs[sebi_dev_cols].head(50),
                                 hide_index=True, use_container_width=True)
                    all_flags.append(("medium", "SEBI Fees",
                                     f"{len(sebi_devs)} trades with SEBI fee deviation"))
                else:
                    st.info(f"All {len(sebi_check):,} trades have SEBI fees at prescribed rate.")
    else:
        st.warning("SEBI fees data not available.")

    with st.expander("Audit Procedures — SEBI Fees"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["SEBI Fees"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 9: STAMP DUTY
# ═══════════════════════════════════════════════════════════════════════════

with tabs[9]:
    section_header("Section 10: Stamp Duty Analysis")

    stamp_col = find_column(df_trades, COL_STAMP_CHARGES)
    if stamp_col:
        total_stamp = df_trades[stamp_col].sum()
        st.markdown(metric_card("Total Stamp Duty (from trades)", format_inr(total_stamp)),
                    unsafe_allow_html=True)

        # Exchange-wise
        if COL_EXCHANGE in df_trades.columns:
            st.markdown("#### Exchange-wise Stamp Duty")
            exch_stamp = df_trades.groupby(COL_EXCHANGE).agg(
                Stamp_Duty=(stamp_col, "sum"),
                Turnover=(COL_TURNOVER, "sum") if COL_TURNOVER in df_trades.columns else (COL_MARKET_VALUE, "sum"),
            ).reset_index()
            exch_stamp["Rate_bps"] = (exch_stamp["Stamp_Duty"] / exch_stamp["Turnover"] * 10000).round(4)
            exch_stamp["Stamp_Duty_fmt"] = exch_stamp["Stamp_Duty"].apply(format_inr)
            st.dataframe(exch_stamp[[COL_EXCHANGE, "Stamp_Duty_fmt", "Rate_bps"]],
                         hide_index=True, use_container_width=True)

        # Monthly
        if "Month" in df_trades.columns:
            st.markdown("#### Monthly Stamp Duty")
            monthly_stamp = df_trades.groupby("Month")[stamp_col].sum().reset_index()
            monthly_stamp["Month"] = monthly_stamp["Month"].astype(str)
            monthly_stamp["Stamp_Duty_fmt"] = monthly_stamp[stamp_col].apply(format_inr)
            st.dataframe(monthly_stamp[["Month", "Stamp_Duty_fmt"]],
                         hide_index=True, use_container_width=True)

        # Cross-check with bank payments
        if df_stamp_bank is not None:
            st.markdown("#### Stamp Duty — Bank Payment Cross-check")
            st.dataframe(df_stamp_bank, use_container_width=True, height=300)
    else:
        st.warning("Stamp duty data not available in trade data.")

    with st.expander("Audit Procedures — Stamp Duty"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["Stamp Duty"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 10: ADJUSTMENTS
# ═══════════════════════════════════════════════════════════════════════════

with tabs[10]:
    section_header("Section 11: Adjustment Analysis")

    if not df_summary.empty:
        adj_rows = df_summary[df_summary["Is_Adj"] == True].copy()
        non_adj = df_summary[df_summary["Is_Adj"] == False].copy()

        total_adj = adj_rows["Gross_Brokerage"].sum() if not adj_rows.empty else 0
        total_gross_summary = non_adj["Gross_Brokerage"].sum()
        adj_pct = safe_div(abs(total_adj), total_gross_summary) * 100

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(metric_card("Total Adjustments", format_inr(total_adj)),
                        unsafe_allow_html=True)
        with c2:
            st.markdown(metric_card("Gross Brokerage (Summary)", format_inr(total_gross_summary)),
                        unsafe_allow_html=True)
        with c3:
            st.markdown(metric_card("Adj as % of Gross", f"{adj_pct:.3f}%"),
                        unsafe_allow_html=True)
        with c4:
            st.markdown(metric_card("Adjustment Count", f"{len(adj_rows)}"),
                        unsafe_allow_html=True)

        if adj_pct > ADJUSTMENT_FLAG_PCT:
            flag_card("high", f"Adjustments ({adj_pct:.2f}%) exceed {ADJUSTMENT_FLAG_PCT}% "
                      "of gross brokerage. Material item requiring separate disclosure.")
            all_flags.append(("high", "Adjustments",
                             f"Adjustments {adj_pct:.2f}% of gross brokerage"))

        if not adj_rows.empty:
            st.markdown("#### Adjustment Details")
            adj_disp = adj_rows[["Segment", "Month", "Gross_Brokerage", "Net_Brokerage"]].copy()
            adj_disp["Amount"] = adj_disp["Gross_Brokerage"].apply(format_inr)
            st.dataframe(adj_disp[["Segment", "Month", "Amount"]],
                         hide_index=True, use_container_width=True)

        # Monthly summary
        st.markdown("#### Monthly Summary from Summary Sheet")
        if not non_adj.empty:
            monthly_summary = non_adj.groupby(["Segment", "Month"]).agg({
                "Gross_Brokerage": "sum",
                "Net_Brokerage": "sum",
                "Net_Turnover_Charges": "sum",
                "Net_Stamp_Duty": "sum",
                "Net_GST": "sum",
                "STT": "sum",
                "SEBI_Fees": "sum",
            }).reset_index()
            for col in ["Gross_Brokerage", "Net_Brokerage"]:
                monthly_summary[f"{col}_fmt"] = monthly_summary[col].apply(format_inr)
            st.dataframe(monthly_summary, hide_index=True, use_container_width=True)
    else:
        st.warning("Summary sheet data not parsed.")

    with st.expander("Audit Procedures — Adjustments"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["Adjustments"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 11: BUY/SELL ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

with tabs[11]:
    section_header("Section 12: Buy/Sell Analysis")

    if COL_BUY_SELL in df_trades.columns and COL_MARKET_VALUE in df_trades.columns:
        df_trades["Side"] = df_trades[COL_BUY_SELL].astype(str).str.strip().str.upper()

        buy_mask = df_trades["Side"].str.startswith("B")
        sell_mask = df_trades["Side"].str.startswith("S")

        buy_value = df_trades.loc[buy_mask, COL_MARKET_VALUE].sum()
        sell_value = df_trades.loc[sell_mask, COL_MARKET_VALUE].sum()
        buy_count = buy_mask.sum()
        sell_count = sell_mask.sum()
        bs_ratio = safe_div(buy_value, sell_value)

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(metric_card("Buy Value", format_inr(buy_value)),
                        unsafe_allow_html=True)
        with c2:
            st.markdown(metric_card("Sell Value", format_inr(sell_value)),
                        unsafe_allow_html=True)
        with c3:
            st.markdown(metric_card("Buy:Sell Ratio", f"{bs_ratio:.2f}"),
                        unsafe_allow_html=True)
        with c4:
            st.markdown(metric_card("Buy / Sell Count", f"{buy_count:,} / {sell_count:,}"),
                        unsafe_allow_html=True)

        # Monthly buy vs sell
        if "Month" in df_trades.columns:
            st.markdown("#### Monthly Buy vs Sell")
            monthly_bs = df_trades.groupby(["Month", "Side"])[COL_MARKET_VALUE].sum().reset_index()
            bs_pivot = monthly_bs.pivot_table(
                index="Month", columns="Side", values=COL_MARKET_VALUE, fill_value=0
            )
            bs_pivot.index = bs_pivot.index.astype(str)
            st.bar_chart(bs_pivot)

        # Clients with large net positions
        if COL_CLIENT_CODE in df_trades.columns:
            st.markdown("#### Clients with Largest Net Positions")
            client_pos = df_trades.copy()
            client_pos["Signed_Value"] = np.where(
                client_pos["Side"].str.startswith("B"),
                client_pos[COL_MARKET_VALUE],
                -client_pos[COL_MARKET_VALUE]
            )
            name_col = COL_CLIENT_NAME if COL_CLIENT_NAME in df_trades.columns else COL_CLIENT_CODE
            net_pos = client_pos.groupby([COL_CLIENT_CODE, name_col])["Signed_Value"].sum().reset_index()
            net_pos["Abs_Position"] = net_pos["Signed_Value"].abs()
            net_pos = net_pos.sort_values("Abs_Position", ascending=False).head(top_n)
            net_pos["Net Position"] = net_pos["Signed_Value"].apply(format_inr)
            net_pos["Direction"] = np.where(net_pos["Signed_Value"] > 0, "Net Buyer", "Net Seller")
            st.dataframe(
                net_pos[[COL_CLIENT_CODE, name_col, "Net Position", "Direction"]],
                hide_index=True, use_container_width=True
            )

        # Delivery vs speculative (same day buy+sell)
        if COL_TXN_DATE in df_trades.columns and COL_SCRIP_CODE in df_trades.columns:
            st.markdown("#### Potential Intraday (Speculative) Trades")
            intraday = df_trades.groupby([COL_CLIENT_CODE, COL_SCRIP_CODE, COL_TXN_DATE, "Side"]).agg(
                Value=(COL_MARKET_VALUE, "sum"),
                Count=(COL_MARKET_VALUE, "count"),
            ).reset_index()
            intraday_pivot = intraday.pivot_table(
                index=[COL_CLIENT_CODE, COL_SCRIP_CODE, COL_TXN_DATE],
                columns="Side", values="Value", fill_value=0
            ).reset_index()

            buy_cols = [c for c in intraday_pivot.columns if str(c).strip().upper() in ("B", "BUY")]
            sell_cols = [c for c in intraday_pivot.columns if str(c).strip().upper() in ("S", "SELL")]
            if buy_cols and sell_cols:
                intraday_pivot["Has_Buy"] = pd.to_numeric(intraday_pivot[buy_cols[0]], errors="coerce").fillna(0) > 0
                intraday_pivot["Has_Sell"] = pd.to_numeric(intraday_pivot[sell_cols[0]], errors="coerce").fillna(0) > 0
                intraday_pairs = intraday_pivot[
                    intraday_pivot["Has_Buy"] & intraday_pivot["Has_Sell"]
                ]
                st.info(f"Detected **{len(intraday_pairs):,}** potential intraday "
                        f"(same client, same scrip, same day buy + sell) instances.")
    else:
        st.warning("Buy/Sell or Market Value data not available.")

    with st.expander("Audit Procedures — Buy/Sell"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["Buy Sell"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 12: TEMPORAL ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

with tabs[12]:
    section_header("Section 13: Temporal Analysis")

    if COL_TXN_DATE in df_trades.columns:
        # Day of week
        st.markdown("#### Day-of-Week Distribution")
        dow = df_trades["Day_of_week"].value_counts()
        dow_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        dow = dow.reindex(dow_order, fill_value=0)
        st.bar_chart(dow)

        weekend_trades = dow.get("Saturday", 0) + dow.get("Sunday", 0)
        if weekend_trades > 0:
            flag_card("high", f"{weekend_trades} trades on weekends detected — data error!")
            all_flags.append(("high", "Temporal",
                             f"{weekend_trades} weekend trades — data error"))

        # Exchange holiday detection
        st.markdown("#### Exchange Holiday Detection (FY 2025-26)")
        holiday_dates = pd.to_datetime(FY2026_HOLIDAYS)
        holiday_mask = df_trades[COL_TXN_DATE].dt.normalize().isin(holiday_dates)
        holiday_trades = df_trades[holiday_mask]
        if not holiday_trades.empty:
            flag_card("high", f"{len(holiday_trades)} trades on exchange holidays detected — data error!")
            all_flags.append(("high", "Temporal",
                             f"{len(holiday_trades)} trades on exchange holidays"))
            hol_summary = holiday_trades.groupby(COL_TXN_DATE).agg(
                Trade_Count=(COL_TOTAL_BROKERAGE, "count"),
                Value=(COL_MARKET_VALUE, "sum") if COL_MARKET_VALUE in holiday_trades.columns else (COL_TOTAL_BROKERAGE, "sum"),
            ).reset_index()
            hol_summary["Value_fmt"] = hol_summary["Value"].apply(format_inr)
            hol_summary["Day"] = hol_summary[COL_TXN_DATE].dt.day_name()
            st.dataframe(hol_summary, hide_index=True, use_container_width=True)
        else:
            st.info("No trades found on exchange holidays.")

        # Combined non-working days summary
        non_working_mask = (
            df_trades["Day_of_week"].isin(["Saturday", "Sunday"]) | holiday_mask
        )
        non_working_total = non_working_mask.sum()
        if non_working_total > 0:
            st.markdown(f"**Total non-working day trades (weekends + holidays): {non_working_total:,}**")

        # Month-end bunching
        st.markdown("#### Month-End Bunching Analysis")
        if "Day_of_month" in df_trades.columns and "Days_in_month" in df_trades.columns:
            df_trades["Is_Month_End"] = (
                df_trades["Days_in_month"] - df_trades["Day_of_month"]
            ) < 5  # Last 5 days

            bunching = df_trades.groupby("Month").agg(
                Total_Trades=("Is_Month_End", "count"),
                Month_End_Trades=("Is_Month_End", "sum"),
            ).reset_index()
            bunching["Month_End_Pct"] = (bunching["Month_End_Trades"] / bunching["Total_Trades"] * 100).round(1)
            bunching["Flag"] = bunching["Month_End_Pct"] > MONTH_END_BUNCHING_PCT
            bunching["Month"] = bunching["Month"].astype(str)
            st.dataframe(bunching, hide_index=True, use_container_width=True)

            flagged = bunching[bunching["Flag"]]
            if not flagged.empty:
                for _, row in flagged.iterrows():
                    flag_card("medium", f"{row['Month']}: {row['Month_End_Pct']:.1f}% of trades "
                              "in last 5 days — month-end bunching")
                all_flags.append(("medium", "Temporal",
                                 f"{len(flagged)} months with month-end bunching"))

        # Volume spikes
        st.markdown("#### Daily Volume Spikes")
        daily = df_trades.groupby(COL_TXN_DATE).agg(
            Trade_Count=(COL_TOTAL_BROKERAGE, "count"),
            Value=(COL_MARKET_VALUE, "sum") if COL_MARKET_VALUE in df_trades.columns else (COL_TOTAL_BROKERAGE, "sum"),
        ).reset_index()

        mean_count = daily["Trade_Count"].mean()
        std_count = daily["Trade_Count"].std()
        daily["Spike"] = daily["Trade_Count"] > (mean_count + VOLUME_SPIKE_STD * std_count)
        spikes = daily[daily["Spike"]].sort_values("Trade_Count", ascending=False)

        if not spikes.empty:
            st.warning(f"{len(spikes)} days with abnormal trading volume detected.")
            spikes["Value_fmt"] = spikes["Value"].apply(format_inr)
            spikes["Std_Multiple"] = ((spikes["Trade_Count"] - mean_count) / std_count).round(1)
            st.dataframe(
                spikes[[COL_TXN_DATE, "Trade_Count", "Value_fmt", "Std_Multiple"]].head(20),
                hide_index=True, use_container_width=True
            )
            all_flags.append(("low", "Temporal",
                             f"{len(spikes)} days with volume spikes"))
        else:
            st.info("No abnormal volume spikes detected.")

        # Daily Turnover Ranking (for OTR file sampling)
        st.markdown("#### Daily Turnover Ranking (for OTR File Sampling)")
        if COL_TURNOVER in df_trades.columns:
            daily_turnover = df_trades.groupby(COL_TXN_DATE).agg(
                Total_Turnover=(COL_TURNOVER, "sum"),
                Trade_Count=(COL_TOTAL_BROKERAGE, "count"),
                Gross_Brokerage=(COL_TOTAL_BROKERAGE, "sum"),
            ).reset_index().sort_values("Total_Turnover", ascending=False)
            daily_turnover["Rank"] = range(1, len(daily_turnover) + 1)
            daily_turnover["Turnover_fmt"] = daily_turnover["Total_Turnover"].apply(format_inr)
            daily_turnover["Brokerage_fmt"] = daily_turnover["Gross_Brokerage"].apply(format_inr)
            daily_turnover["Day"] = daily_turnover[COL_TXN_DATE].dt.day_name()
            st.info(f"Total trading days: {len(daily_turnover)}")
            st.dataframe(
                daily_turnover[["Rank", COL_TXN_DATE, "Day", "Trade_Count",
                                "Turnover_fmt", "Brokerage_fmt"]].head(top_n),
                hide_index=True, use_container_width=True
            )
    else:
        st.warning("Transaction date not available for temporal analysis.")

    with st.expander("Audit Procedures — Temporal"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES["Temporal"]:
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 13: CLIENT × PRODUCT ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

with tabs[13]:
    section_header("Section 14: Client × Product Analysis")

    if COL_CLIENT_CODE in df_trades.columns and COL_TOTAL_BROKERAGE in df_trades.columns and "Segment" in df_trades.columns:
        name_col_cp = COL_CLIENT_NAME if COL_CLIENT_NAME in df_trades.columns else COL_CLIENT_CODE

        # Client × Segment aggregation
        cp_agg = {
            "Trade_Count": (COL_TOTAL_BROKERAGE, "count"),
            "Gross_Brokerage": (COL_TOTAL_BROKERAGE, "sum"),
        }
        if COL_TURNOVER in df_trades.columns:
            cp_agg["Turnover"] = (COL_TURNOVER, "sum")
        if COL_NET_BROKERAGE in df_trades.columns:
            cp_agg["Net_Brokerage"] = (COL_NET_BROKERAGE, "sum")

        client_product = df_trades.groupby(
            [COL_CLIENT_CODE, name_col_cp, "Segment"]
        ).agg(**cp_agg).reset_index()

        # Total per client (for ranking and dominance)
        client_total_cp = client_product.groupby([COL_CLIENT_CODE, name_col_cp]).agg(
            Total_Brokerage=("Gross_Brokerage", "sum"),
            Total_Trades=("Trade_Count", "sum"),
        ).reset_index().sort_values("Total_Brokerage", ascending=False)

        top_client_codes = client_total_cp.head(top_n)[COL_CLIENT_CODE].tolist()

        # Multi-segment metrics
        seg_per_client = client_product.groupby(COL_CLIENT_CODE)["Segment"].nunique()
        multi_seg_count = int((seg_per_client > 1).sum())
        single_seg_count = int((seg_per_client == 1).sum())

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(metric_card("Total Clients", f"{len(client_total_cp):,}"),
                        unsafe_allow_html=True)
        with c2:
            st.markdown(metric_card("Multi-Segment Clients", f"{multi_seg_count:,}"),
                        unsafe_allow_html=True)
        with c3:
            st.markdown(metric_card("Single-Segment Clients", f"{single_seg_count:,}"),
                        unsafe_allow_html=True)
        with c4:
            seg_cnt = client_product["Segment"].nunique()
            st.markdown(metric_card("Product Segments", f"{seg_cnt}"),
                        unsafe_allow_html=True)

        # ── Brokerage pivot ──
        st.markdown(f"#### Top {top_n} Clients — Brokerage by Product Segment")
        brok_pivot = client_product[
            client_product[COL_CLIENT_CODE].isin(top_client_codes)
        ].pivot_table(
            index=[COL_CLIENT_CODE, name_col_cp],
            columns="Segment", values="Gross_Brokerage",
            aggfunc="sum", fill_value=0,
        )
        brok_pivot["Total"] = brok_pivot.sum(axis=1)
        brok_pivot = brok_pivot.sort_values("Total", ascending=False)
        brok_disp = brok_pivot.copy()
        for col in brok_disp.columns:
            brok_disp[col] = brok_disp[col].apply(format_inr)
        st.dataframe(brok_disp, use_container_width=True)

        # ── Trade count pivot ──
        st.markdown(f"#### Top {top_n} Clients — Trade Count by Segment")
        count_pivot = client_product[
            client_product[COL_CLIENT_CODE].isin(top_client_codes)
        ].pivot_table(
            index=[COL_CLIENT_CODE, name_col_cp],
            columns="Segment", values="Trade_Count",
            aggfunc="sum", fill_value=0,
        )
        count_pivot["Total"] = count_pivot.sum(axis=1)
        count_pivot = count_pivot.sort_values("Total", ascending=False)
        st.dataframe(count_pivot, use_container_width=True)

        # ── Brokerage rate pivot ──
        if COL_TURNOVER in df_trades.columns:
            st.markdown(f"#### Top {top_n} Clients — Avg Brokerage Rate (bps) by Segment")
            rate_data = client_product[
                client_product[COL_CLIENT_CODE].isin(top_client_codes)
            ].copy()
            rate_data["Rate_bps"] = np.where(
                rate_data["Turnover"] > 0,
                rate_data["Gross_Brokerage"] / rate_data["Turnover"] * 10000, 0
            ).round(2)
            rate_piv = rate_data.pivot_table(
                index=[COL_CLIENT_CODE, name_col_cp],
                columns="Segment", values="Rate_bps",
                aggfunc="first", fill_value=0,
            )
            st.dataframe(rate_piv, use_container_width=True)

        # ── Segment dominance ──
        st.markdown("#### Client Segment Dominance")
        dominance = client_product.merge(
            client_total_cp[[COL_CLIENT_CODE, "Total_Brokerage"]], on=COL_CLIENT_CODE
        )
        dominance["Segment_Pct"] = (
            dominance["Gross_Brokerage"] / dominance["Total_Brokerage"] * 100
        ).round(1)

        # Clients >90% from single segment
        dominant = dominance[dominance["Segment_Pct"] > 90]
        dom_summary = dominant.groupby("Segment")[COL_CLIENT_CODE].nunique().reset_index()
        dom_summary.columns = ["Segment", "Clients with >90% in Segment"]
        st.dataframe(dom_summary, hide_index=True, use_container_width=True)

        # Segment revenue contribution
        st.markdown("#### Revenue Mix by Product Segment")
        seg_rev = client_product.groupby("Segment").agg(
            Gross_Brokerage=("Gross_Brokerage", "sum"),
            Clients=(COL_CLIENT_CODE, "nunique"),
            Trades=("Trade_Count", "sum"),
        ).reset_index()
        seg_rev["% of Revenue"] = (
            seg_rev["Gross_Brokerage"] / seg_rev["Gross_Brokerage"].sum() * 100
        ).round(2)
        seg_rev["Brokerage (Formatted)"] = seg_rev["Gross_Brokerage"].apply(format_inr)
        st.dataframe(
            seg_rev[["Segment", "Clients", "Trades", "Brokerage (Formatted)", "% of Revenue"]],
            hide_index=True, use_container_width=True
        )
        single_seg_dominant = seg_rev.loc[seg_rev["% of Revenue"].idxmax()]
        if single_seg_dominant["% of Revenue"] > 70:
            flag_card("medium",
                      f"{single_seg_dominant['Segment']} contributes "
                      f"{single_seg_dominant['% of Revenue']:.1f}% of total revenue — "
                      "segment concentration risk.")
            all_flags.append(("medium", "Client × Product",
                             f"{single_seg_dominant['Segment']} = "
                             f"{single_seg_dominant['% of Revenue']:.1f}% of revenue"))

        # ── Product Type Rate Uniformity Check ──
        if COL_PRODUCT_DESC in df_trades.columns:
            st.markdown("#### Product Type Rate Uniformity Check")
            st.caption("Verifies Brokerage Rate % = Gross Brokerage / Market Value is uniform "
                       "within each Client × Product Type combination.")

            prod_check = df_trades[
                (df_trades[COL_TOTAL_BROKERAGE] > 0) & (df_trades[COL_MARKET_VALUE] > 0)
            ].copy()
            prod_check["Brok_Rate_MV"] = (
                prod_check[COL_TOTAL_BROKERAGE] / prod_check[COL_MARKET_VALUE]
            )

            rate_unif = prod_check.groupby(
                [COL_CLIENT_CODE, name_col_cp, COL_PRODUCT_DESC]
            ).agg(
                Trade_Count=("Brok_Rate_MV", "count"),
                Min_Rate=("Brok_Rate_MV", "min"),
                Max_Rate=("Brok_Rate_MV", "max"),
                Mean_Rate=("Brok_Rate_MV", "mean"),
                Std_Rate=("Brok_Rate_MV", "std"),
                Total_Brokerage=(COL_TOTAL_BROKERAGE, "sum"),
            ).reset_index()
            rate_unif["Std_Rate"] = rate_unif["Std_Rate"].fillna(0)
            rate_unif["Range_bps"] = (
                (rate_unif["Max_Rate"] - rate_unif["Min_Rate"]) * 10000
            ).round(2)

            non_uniform = rate_unif[
                (rate_unif["Range_bps"] > 1) & (rate_unif["Trade_Count"] > 1)
            ].sort_values("Range_bps", ascending=False)

            if not non_uniform.empty:
                st.warning(f"{len(non_uniform)} Client × Product combinations with non-uniform "
                           "brokerage rates.")
                nu_disp = non_uniform.copy()
                for rc in ["Min_Rate", "Max_Rate", "Mean_Rate"]:
                    nu_disp[f"{rc}_bps"] = (nu_disp[rc] * 10000).round(2)
                nu_disp["Brokerage"] = nu_disp["Total_Brokerage"].apply(format_inr)
                st.dataframe(
                    nu_disp[[COL_CLIENT_CODE, name_col_cp, COL_PRODUCT_DESC,
                             "Trade_Count", "Min_Rate_bps", "Max_Rate_bps",
                             "Mean_Rate_bps", "Range_bps", "Brokerage"]].head(50),
                    hide_index=True, use_container_width=True
                )
                all_flags.append(("medium", "Client × Product",
                                 f"{len(non_uniform)} client-product combos with non-uniform rates"))
            else:
                st.info("All Client × Product Type combinations have uniform brokerage rates.")

            # Product Type Distribution
            st.markdown("#### Product Type Distribution")
            prod_dist = df_trades.groupby(COL_PRODUCT_DESC).agg(
                Trade_Count=(COL_TOTAL_BROKERAGE, "count"),
                Gross_Brokerage=(COL_TOTAL_BROKERAGE, "sum"),
                Turnover_Total=(COL_TURNOVER, "sum") if COL_TURNOVER in df_trades.columns else (COL_TOTAL_BROKERAGE, "sum"),
            ).reset_index().sort_values("Gross_Brokerage", ascending=False)
            prod_dist["Avg_Rate_bps"] = np.where(
                prod_dist["Turnover_Total"] > 0,
                prod_dist["Gross_Brokerage"] / prod_dist["Turnover_Total"] * 10000, 0
            ).round(2)
            prod_dist["% of Brokerage"] = (
                prod_dist["Gross_Brokerage"] / prod_dist["Gross_Brokerage"].sum() * 100
            ).round(2)
            prod_dist["Brokerage_fmt"] = prod_dist["Gross_Brokerage"].apply(format_inr)
            st.dataframe(
                prod_dist[[COL_PRODUCT_DESC, "Trade_Count", "Brokerage_fmt",
                           "Avg_Rate_bps", "% of Brokerage"]],
                hide_index=True, use_container_width=True
            )

        # ── Client search ──
        st.markdown("#### Client-wise Product Detail (Search)")
        search_cp = st.text_input("Search Client (Code or Name)", key="cp_search")
        if search_cp:
            mask_cp = (
                client_product[COL_CLIENT_CODE].astype(str).str.contains(
                    search_cp, case=False, na=False) |
                client_product[name_col_cp].astype(str).str.contains(
                    search_cp, case=False, na=False)
            )
            filtered_cp = client_product[mask_cp].copy()
            if not filtered_cp.empty:
                for col_f in ["Gross_Brokerage"]:
                    if col_f in filtered_cp.columns:
                        filtered_cp[f"{col_f}_fmt"] = filtered_cp[col_f].apply(format_inr)
                if "Turnover" in filtered_cp.columns:
                    filtered_cp["Turnover_fmt"] = filtered_cp["Turnover"].apply(format_inr)
                    filtered_cp["Rate (bps)"] = np.where(
                        filtered_cp["Turnover"] > 0,
                        filtered_cp["Gross_Brokerage"] / filtered_cp["Turnover"] * 10000, 0
                    ).round(2)
                st.dataframe(filtered_cp, hide_index=True, use_container_width=True)
            else:
                st.info("No clients found matching your search.")
    else:
        st.warning("Client code, brokerage, or segment data not available.")

    with st.expander("Audit Procedures — Client × Product Analysis"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES.get("Client Product", []):
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 14: TRANSACTION-LEVEL ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

with tabs[14]:
    section_header("Section 15: Transaction-Level Analysis")

    if not df_trades.empty:
        # ── Filters ──
        st.markdown("#### Filters")
        fc1, fc2, fc3, fc4 = st.columns(4)
        with fc1:
            seg_opts = df_trades["Segment"].unique().tolist()
            seg_filter = st.multiselect("Segment", seg_opts, default=seg_opts, key="txn_seg")
        with fc2:
            if COL_EXCHANGE in df_trades.columns:
                exch_opts = df_trades[COL_EXCHANGE].dropna().unique().tolist()
                exch_filter = st.multiselect("Exchange", exch_opts, default=exch_opts,
                                              key="txn_exch")
            else:
                exch_filter = None
        with fc3:
            if COL_TXN_DATE in df_trades.columns:
                min_dt = df_trades[COL_TXN_DATE].min()
                max_dt = df_trades[COL_TXN_DATE].max()
                if pd.notna(min_dt) and pd.notna(max_dt):
                    date_range = st.date_input("Date Range",
                                                value=(min_dt.date(), max_dt.date()),
                                                min_value=min_dt.date(),
                                                max_value=max_dt.date(),
                                                key="txn_date")
                else:
                    date_range = None
            else:
                date_range = None
        with fc4:
            if COL_BUY_SELL in df_trades.columns:
                side_opts = df_trades[COL_BUY_SELL].dropna().unique().tolist()
                side_filter = st.multiselect("Buy/Sell", side_opts, default=side_opts,
                                              key="txn_side")
            else:
                side_filter = None

        # Apply filters
        filt = df_trades[df_trades["Segment"].isin(seg_filter)].copy()
        if exch_filter is not None and COL_EXCHANGE in filt.columns:
            filt = filt[filt[COL_EXCHANGE].isin(exch_filter)]
        if date_range is not None and len(date_range) == 2 and COL_TXN_DATE in filt.columns:
            filt = filt[
                (filt[COL_TXN_DATE] >= pd.Timestamp(date_range[0])) &
                (filt[COL_TXN_DATE] <= pd.Timestamp(date_range[1]))
            ]
        if side_filter is not None and COL_BUY_SELL in filt.columns:
            filt = filt[filt[COL_BUY_SELL].isin(side_filter)]

        f_total = len(filt)
        f_brok = filt[COL_TOTAL_BROKERAGE].sum() if COL_TOTAL_BROKERAGE in filt.columns else 0
        f_turnover = filt[COL_TURNOVER].sum() if COL_TURNOVER in filt.columns else 0
        f_avg_rate = safe_div(f_brok, f_turnover) * 10000

        # ── Summary metrics ──
        st.markdown("#### Filtered Dataset Summary")
        mc1, mc2, mc3, mc4 = st.columns(4)
        with mc1:
            st.markdown(metric_card("Filtered Trades", f"{f_total:,}"),
                        unsafe_allow_html=True)
        with mc2:
            st.markdown(metric_card("Total Brokerage", format_inr(f_brok)),
                        unsafe_allow_html=True)
        with mc3:
            st.markdown(metric_card("Total Turnover", format_inr(f_turnover)),
                        unsafe_allow_html=True)
        with mc4:
            st.markdown(metric_card("Avg Rate (bps)", f"{f_avg_rate:.2f}"),
                        unsafe_allow_html=True)

        # ── Distribution statistics ──
        st.markdown("#### Brokerage Distribution Statistics")
        if COL_TOTAL_BROKERAGE in filt.columns and f_total > 0:
            bs = filt[COL_TOTAL_BROKERAGE]
            dist_data = pd.DataFrame({
                "Statistic": ["Count", "Mean", "Median", "Std Dev", "Min",
                              "25th Percentile", "75th Percentile", "Max",
                              "Skewness", "Kurtosis"],
                "Brokerage (Rs.)": [
                    f_total, round(bs.mean(), 2), round(bs.median(), 2),
                    round(bs.std(), 2), round(bs.min(), 2),
                    round(bs.quantile(0.25), 2), round(bs.quantile(0.75), 2),
                    round(bs.max(), 2), round(bs.skew(), 2), round(bs.kurtosis(), 2),
                ],
            })
            st.dataframe(dist_data, hide_index=True, use_container_width=True)

        # ── Display columns for trade tables ──
        txn_disp_cols = [c for c in [COL_TXN_DATE, COL_CLIENT_CODE, COL_CLIENT_NAME,
                                      COL_SCRIP_NAME, "Segment", COL_BUY_SELL,
                                      COL_TURNOVER, COL_TOTAL_BROKERAGE, "Brokerage_Rate"]
                         if c in filt.columns]

        # ── Top trades by brokerage ──
        st.markdown(f"#### Top {top_n} Trades by Brokerage Amount")
        if COL_TOTAL_BROKERAGE in filt.columns:
            top_brok = filt.nlargest(top_n, COL_TOTAL_BROKERAGE)[txn_disp_cols].copy()
            if "Brokerage_Rate" in top_brok.columns:
                top_brok["Rate (bps)"] = (top_brok["Brokerage_Rate"] * 10000).round(2)
            st.dataframe(top_brok, hide_index=True, use_container_width=True)

        # ── Top trades by turnover ──
        if COL_TURNOVER in filt.columns:
            st.markdown(f"#### Top {top_n} Trades by Turnover")
            top_tv = filt.nlargest(top_n, COL_TURNOVER)[txn_disp_cols].copy()
            if "Brokerage_Rate" in top_tv.columns:
                top_tv["Rate (bps)"] = (top_tv["Brokerage_Rate"] * 10000).round(2)
            st.dataframe(top_tv, hide_index=True, use_container_width=True)

        # ── Anomalous trades ──
        st.markdown("#### Anomalous Trades")
        if COL_TOTAL_BROKERAGE in filt.columns:
            zero_brok = filt[filt[COL_TOTAL_BROKERAGE] == 0]
            neg_brok = filt[filt[COL_TOTAL_BROKERAGE] < 0]

            ac1, ac2 = st.columns(2)
            with ac1:
                st.markdown(metric_card("Zero Brokerage Trades", f"{len(zero_brok):,}"),
                            unsafe_allow_html=True)
            with ac2:
                st.markdown(metric_card("Negative Brokerage Trades", f"{len(neg_brok):,}"),
                            unsafe_allow_html=True)

            if not neg_brok.empty:
                flag_card("high", f"{len(neg_brok)} trades with negative brokerage — "
                          "verify if these are reversals or data errors.")
                all_flags.append(("high", "Txn Analysis",
                                 f"{len(neg_brok)} trades with negative brokerage"))
                st.dataframe(neg_brok[txn_disp_cols].head(30),
                             hide_index=True, use_container_width=True)

            if len(zero_brok) > 0:
                zero_pct = len(zero_brok) / max(f_total, 1) * 100
                sev_zero = "medium" if zero_pct > 5 else "low"
                flag_card(sev_zero,
                          f"{len(zero_brok):,} trades ({zero_pct:.1f}%) with zero brokerage — "
                          "verify if proprietary, error, or promotional trades.")
                all_flags.append((sev_zero, "Txn Analysis",
                                 f"{len(zero_brok):,} ({zero_pct:.1f}%) zero brokerage trades"))
                with st.expander(f"View Zero Brokerage Trades (up to 100)"):
                    st.dataframe(zero_brok[txn_disp_cols].head(100),
                                 hide_index=True, use_container_width=True)

        # ── SEBI Maximum 2.5% Brokerage Check ──
        st.markdown("#### SEBI Maximum Brokerage Rate (2.5%) Check")
        if "Brokerage_Rate" in filt.columns:
            over_sebi = filt[filt["Brokerage_Rate"] > SEBI_MAX_BROKERAGE_PCT]
            if not over_sebi.empty:
                flag_card("high", f"{len(over_sebi)} trades exceed SEBI maximum brokerage "
                          "rate of 2.5%. Excess brokerage must be refunded to clients.")
                all_flags.append(("high", "Txn Analysis",
                                 f"{len(over_sebi)} trades exceed SEBI max 2.5% brokerage"))
                over_disp = over_sebi[txn_disp_cols].head(50).copy()
                if "Brokerage_Rate" in over_disp.columns:
                    over_disp["Rate %"] = (over_disp["Brokerage_Rate"] * 100).round(4)
                st.dataframe(over_disp, hide_index=True, use_container_width=True)
            else:
                st.info("No trades exceed the SEBI maximum brokerage rate of 2.5%.")

        # ── High Turnover / Low Brokerage ──
        st.markdown("#### High Turnover with Low Brokerage")
        if COL_TURNOVER in filt.columns and "Brokerage_Rate" in filt.columns:
            valid_htlb = filt[(filt[COL_TURNOVER] > 0) & (filt["Brokerage_Rate"] > 0)].copy()
            if len(valid_htlb) > 20:
                tv_75 = valid_htlb[COL_TURNOVER].quantile(0.75)
                rate_10 = valid_htlb["Brokerage_Rate"].quantile(0.10)
                htlb = valid_htlb[
                    (valid_htlb[COL_TURNOVER] >= tv_75) &
                    (valid_htlb["Brokerage_Rate"] <= rate_10)
                ]
                if not htlb.empty:
                    flag_card("medium", f"{len(htlb)} trades in top 25% turnover but bottom "
                              "10% brokerage rate — potential revenue leakage or preferential rates.")
                    all_flags.append(("medium", "Txn Analysis",
                                     f"{len(htlb)} high turnover / low brokerage trades"))
                    htlb_disp = htlb[txn_disp_cols].head(50).copy()
                    if "Brokerage_Rate" in htlb_disp.columns:
                        htlb_disp["Rate (bps)"] = (htlb_disp["Brokerage_Rate"] * 10000).round(2)
                    st.dataframe(htlb_disp, hide_index=True, use_container_width=True)
                else:
                    st.info("No trades with high turnover and disproportionately low brokerage.")
            else:
                st.info("Insufficient trades for high turnover / low brokerage analysis.")

        # ── Trades above materiality ──
        if COL_TOTAL_BROKERAGE in filt.columns:
            st.markdown("#### Trades Above Materiality")
            material_trades = filt[filt[COL_TOTAL_BROKERAGE].abs() >= materiality]
            st.info(f"**{len(material_trades):,}** trades with brokerage >= "
                    f"Rs. {materiality:,.0f}")
            if not material_trades.empty:
                st.dataframe(material_trades[txn_disp_cols].head(50),
                             hide_index=True, use_container_width=True)

        # ── Benford's Law ──
        st.markdown("#### Benford's Law — First Digit Analysis")
        if COL_TOTAL_BROKERAGE in filt.columns:
            brok_pos = filt[COL_TOTAL_BROKERAGE][filt[COL_TOTAL_BROKERAGE] > 0]
            if len(brok_pos) > 100:
                first_dig = brok_pos.apply(
                    lambda x: int(str(x).lstrip("0").replace(".", "")[0])
                    if x > 0 else 0
                )
                first_dig = first_dig[first_dig > 0]
                observed = first_dig.value_counts(normalize=True).sort_index()
                expected_bf = pd.Series(
                    {d: math.log10(1 + 1 / d) for d in range(1, 10)}
                )
                benford_df = pd.DataFrame({
                    "Digit": range(1, 10),
                    "Expected %": (expected_bf.values * 100).round(2),
                    "Observed %": [round(observed.get(d, 0) * 100, 2) for d in range(1, 10)],
                })
                benford_df["Deviation %"] = (
                    benford_df["Observed %"] - benford_df["Expected %"]
                ).round(2)
                st.dataframe(benford_df, hide_index=True, use_container_width=True)

                st.bar_chart(benford_df.set_index("Digit")[["Expected %", "Observed %"]])

                mad_val = benford_df["Deviation %"].abs().mean()
                if mad_val > 1.5:
                    flag_card("medium",
                              f"Benford's Law MAD: {mad_val:.2f}% — significant deviation, "
                              "investigate brokerage computation patterns.")
                    all_flags.append(("medium", "Txn Analysis",
                                     f"Benford's Law MAD {mad_val:.2f}%"))
                else:
                    st.info(f"Benford's Law test: MAD = {mad_val:.2f}% (conforming)")
            else:
                st.info("Insufficient data for Benford's Law test (need > 100 positive trades).")

        # ── Audit sample selection ──
        st.markdown("#### Audit Sample Selection (SA 530)")
        sc1, sc2 = st.columns(2)
        with sc1:
            sample_size = st.number_input("Sample Size", value=30, step=5,
                                           min_value=5, max_value=200, key="txn_sample")
        with sc2:
            sample_method = st.radio("Sampling Method",
                                      ["Random", "Monetary Unit (MUS)", "Stratified"],
                                      horizontal=True, key="txn_method")

        if st.button("Generate Sample", key="txn_gen_sample"):
            n_sample = min(sample_size, len(filt))
            if sample_method == "Random":
                sample_df = filt.sample(n=n_sample, random_state=42)
            elif sample_method == "Monetary Unit (MUS)":
                if COL_TOTAL_BROKERAGE in filt.columns:
                    w = filt[COL_TOTAL_BROKERAGE].abs().replace(0, 0.01)
                    w = w / w.sum()
                    sample_df = filt.sample(n=n_sample, weights=w, random_state=42)
                else:
                    sample_df = filt.sample(n=n_sample, random_state=42)
            else:
                per_seg = max(1, sample_size // max(1, filt["Segment"].nunique()))
                frames = []
                for seg_name, seg_df in filt.groupby("Segment"):
                    frames.append(seg_df.sample(
                        n=min(per_seg, len(seg_df)), random_state=42
                    ))
                sample_df = pd.concat(frames).reset_index(drop=True)

            st.markdown(f"**{len(sample_df)} trades selected — {sample_method} sampling**")
            st.dataframe(sample_df[txn_disp_cols], hide_index=True, use_container_width=True)
    else:
        st.warning("No trade data available for transaction-level analysis.")

    with st.expander("Audit Procedures — Transaction Analysis"):
        for step_no, procedure, evidence in AUDIT_PROCEDURES.get("Transaction Analysis", []):
            st.markdown(f"**Step {step_no}:** {procedure}")
            st.markdown(f"*Expected Evidence:* {evidence}")
            st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════
# TAB 15: AUDIT FLAGS CONSOLIDATED
# ═══════════════════════════════════════════════════════════════════════════

with tabs[15]:
    section_header("Consolidated Audit Flags")

    if all_flags:
        high_flags = [f for f in all_flags if f[0] == "high"]
        med_flags = [f for f in all_flags if f[0] == "medium"]
        low_flags = [f for f in all_flags if f[0] == "low"]

        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(metric_card("High Severity", f"{len(high_flags)}"),
                        unsafe_allow_html=True)
        with c2:
            st.markdown(metric_card("Medium Severity", f"{len(med_flags)}"),
                        unsafe_allow_html=True)
        with c3:
            st.markdown(metric_card("Low Severity", f"{len(low_flags)}"),
                        unsafe_allow_html=True)

        if high_flags:
            st.markdown("#### High Severity")
            for severity, section, msg in high_flags:
                flag_card("high", f"**[{section}]** {msg}")

        if med_flags:
            st.markdown("#### Medium Severity")
            for severity, section, msg in med_flags:
                flag_card("medium", f"**[{section}]** {msg}")

        if low_flags:
            st.markdown("#### Low Severity")
            for severity, section, msg in low_flags:
                flag_card("low", f"**[{section}]** {msg}")
    else:
        st.info("No audit flags raised. All checks passed.")


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL REPORT GENERATION
# ═══════════════════════════════════════════════════════════════════════════

st.markdown("---")
section_header("Export Report")


def generate_excel_report():
    """Generate the comprehensive Excel report with KKC branding."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── KKC Styles ──
    hdr_fill = PatternFill(start_color=KKC_GREEN, end_color=KKC_GREEN, fill_type="solid")
    hdr_font = Font(name=FONT_NAME, bold=True, color=WHITE, size=11)
    title_font = Font(name=FONT_NAME, bold=True, color=DARK_TEXT, size=16)
    subtitle_font = Font(name=FONT_NAME, bold=False, color=KKC_GREY, size=12)
    sec_font = Font(name=FONT_NAME, bold=True, color=KKC_GREEN, size=13)
    body = Font(name=FONT_NAME, size=10)
    body_b = Font(name=FONT_NAME, bold=True, size=10)
    total_font = Font(name=FONT_NAME, bold=True, size=10, color=KKC_GREEN)
    note_font = Font(name=FONT_NAME, italic=True, color=KKC_GREY, size=9)
    bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    thick_bottom = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("medium"))
    wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ctr = Alignment(horizontal="center", vertical="center")
    rt = Alignment(horizontal="right", vertical="center")
    alt_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
    proc_hdr_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    hi_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
    med_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
    lo_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    INR_FMT = '#,##0.00'
    PCT_FMT = '0.000"%"'
    BPS_FMT = '0.0000'
    INT_FMT = '#,##0'

    # ── Reusable writers ──
    def hdr_row(ws, r, vals, col=1):
        for i, v in enumerate(vals):
            c = ws.cell(r, col + i, v)
            c.font, c.fill, c.border, c.alignment = hdr_font, hdr_fill, bdr, ctr

    def data_row(ws, r, vals, col=1, bold=False, stripe=False):
        for i, v in enumerate(vals):
            c = ws.cell(r, col + i, v)
            c.font = body_b if bold else body
            c.border = bdr
            if stripe:
                c.fill = alt_fill
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                c.alignment = rt
                if isinstance(v, int):
                    c.number_format = INT_FMT
                else:
                    c.number_format = INR_FMT
            else:
                c.alignment = wrap

    def total_row(ws, r, vals, col=1):
        for i, v in enumerate(vals):
            c = ws.cell(r, col + i, v)
            c.font = total_font
            c.border = thick_bottom
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                c.alignment = rt
                c.number_format = INR_FMT if isinstance(v, float) else INT_FMT
            else:
                c.alignment = wrap

    def sec_title(ws, r, text, col=1):
        c = ws.cell(r, col, text)
        c.font = sec_font

    def auto_w(ws, mx=55):
        for col_cells in ws.columns:
            ml = 10
            cl = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    ml = max(ml, min(len(str(cell.value)), mx))
            ws.column_dimensions[cl].width = ml + 3

    def sheet_header(ws, title_text, subtitle_text=None):
        """Write standard sheet header: title row 1, subtitle row 2, entity row 3."""
        ws.sheet_properties.tabColor = KKC_GREEN
        ws.cell(1, 1, title_text).font = title_font
        if subtitle_text:
            ws.cell(2, 1, subtitle_text).font = subtitle_font
        else:
            ws.cell(2, 1, f"{entity_name} | {audit_period}").font = subtitle_font

    def write_df(ws, df, start, section=None, show_total_cols=None):
        """Write DataFrame with headers, alternating rows, optional totals, and procedures."""
        if df.empty:
            ws.cell(start, 1, "No data available.").font = note_font
            r = start + 1
        else:
            hdr_row(ws, start, list(df.columns))
            r = start + 1
            for idx, (_, row_data) in enumerate(df.iterrows()):
                data_row(ws, r, list(row_data.values), stripe=(idx % 2 == 1))
                r += 1
            # Total row
            if show_total_cols and len(df) > 1:
                tot_vals = []
                for c_name in df.columns:
                    if c_name in show_total_cols:
                        tot_vals.append(df[c_name].sum())
                    elif df.columns.get_loc(c_name) == 0:
                        tot_vals.append("TOTAL")
                    else:
                        tot_vals.append("")
                total_row(ws, r, tot_vals)
                r += 1
        # Audit procedures
        if section and section in AUDIT_PROCEDURES:
            r += 2
            sec_title(ws, r, f"Audit Procedures - {section}")
            r += 1
            hdr_row(ws, r, ["Step", "Procedure", "Expected Evidence", "Auditor Remarks"])
            r += 1
            for sno, proc, ev in AUDIT_PROCEDURES[section]:
                data_row(ws, r, [sno, proc, ev, ""])
                ws.cell(r, 2).alignment = wrap
                ws.cell(r, 3).alignment = wrap
                ws.cell(r, 4).alignment = wrap
                r += 1
        auto_w(ws)
        return r

    def kv_block(ws, start, pairs):
        """Write key-value pairs as a 2-column block with borders."""
        hdr_row(ws, start, ["Metric", "Value"])
        for i, (k, v) in enumerate(pairs):
            r = start + 1 + i
            ws.cell(r, 1, k).font = body_b
            ws.cell(r, 1).border = bdr
            ws.cell(r, 2, v).font = body
            ws.cell(r, 2).border = bdr
            if isinstance(v, (int, float)):
                ws.cell(r, 2).alignment = rt
                ws.cell(r, 2).number_format = INR_FMT if isinstance(v, float) else INT_FMT
        return start + 1 + len(pairs) + 1

    # ══════════════════════════════════════════════════════
    # Sheet 1: Cover
    # ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = KKC_GREEN
    # Green banner
    for c in range(1, 8):
        ws.cell(2, c).fill = hdr_fill
        ws.cell(3, c).fill = hdr_fill
    ws.merge_cells("B2:G2")
    ws.cell(2, 2, entity_name).font = Font(name=FONT_NAME, bold=True, color=WHITE, size=20)
    ws.cell(2, 2).alignment = Alignment(vertical="center")
    ws.merge_cells("B3:G3")
    ws.cell(3, 2, f"Brokerage Data Analytics - {audit_period}").font = Font(name=FONT_NAME, color=WHITE, size=13)
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 25

    ws.cell(5, 2, "Generated").font = body_b
    ws.cell(5, 3, datetime.now().strftime("%d %B %Y, %H:%M")).font = body
    ws.cell(6, 2, "Tool Version").font = body_b
    ws.cell(6, 3, APP_FULL_NAME).font = body
    ws.cell(7, 2, "Prepared by").font = body_b
    ws.cell(7, 3, "KKC & Associates LLP").font = Font(name=FONT_NAME, bold=True, color=KKC_GREEN, size=11)
    ws.cell(8, 2, "").font = body
    ws.cell(8, 3, "Chartered Accountants, Mumbai").font = subtitle_font

    # Key numbers
    sec_title(ws, 10, "Key Figures")
    kv_block(ws, 11, [
        ("Total Gross Brokerage (Rs.)", round(total_gross, 2)),
        ("Total Net Brokerage (Rs.)", round(total_net, 2)),
        ("Total Turnover (Rs.)", round(total_turnover, 2)),
        ("Total Trades", total_trades),
        ("Total Error Trades", total_errors),
        ("Error Rate (%)", round(error_rate, 3)),
        ("Net Brokerage Margin (%)", round(net_margin, 1)),
    ])

    ws.cell(20, 2,
            "This report has been generated using automated analytics. "
            "All observations must be verified by the engagement team "
            "before forming audit conclusions.").font = note_font
    ws.cell(20, 2).alignment = wrap
    ws.merge_cells("B20:G20")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    for c in "DEFG":
        ws.column_dimensions[c].width = 15

    # ══════════════════════════════════════════════════════
    # Sheet 2: Executive Summary
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Executive Summary")
    sheet_header(ws2, "Executive Summary")

    r = 4
    sec_title(ws2, r, "Audit Flags Summary")
    r += 1
    if all_flags:
        hdr_row(ws2, r, ["Severity", "Section", "Observation"])
        r += 1
        for severity, section, msg in all_flags:
            data_row(ws2, r, [severity.upper(), section, msg])
            fill = hi_fill if severity == "high" else (med_fill if severity == "medium" else lo_fill)
            for c in range(1, 4):
                ws2.cell(r, c).fill = fill
            r += 1
    else:
        ws2.cell(r, 1, "No audit flags raised. All checks passed.").font = note_font
        r += 1

    r += 1
    sec_title(ws2, r, "Segment-wise Brokerage Summary")
    r += 1
    if "Segment" in df_trades.columns:
        seg_df = df_trades.groupby("Segment").agg(
            Gross_Brokerage=(COL_TOTAL_BROKERAGE, "sum"),
            Trade_Count=(COL_TOTAL_BROKERAGE, "count"),
        ).reset_index()
        seg_df["Pct"] = (seg_df["Gross_Brokerage"] / seg_df["Gross_Brokerage"].sum() * 100).round(2)
        seg_df.columns = ["Segment", "Gross Brokerage (Rs.)", "Trade Count", "% of Total"]
        write_df(ws2, seg_df, r, show_total_cols=["Gross Brokerage (Rs.)", "Trade Count"])
    auto_w(ws2)

    # ══════════════════════════════════════════════════════
    # Sheet 3: Summary Dashboard
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Summary Dashboard")
    sheet_header(ws3, "Monthly Brokerage Summary")

    if "Month" in df_trades.columns:
        me_agg = {"Gross Brokerage (Rs.)": (COL_TOTAL_BROKERAGE, "sum"), "Trade Count": (COL_TOTAL_BROKERAGE, "count")}
        if COL_NET_BROKERAGE in df_trades.columns:
            me_agg["Net Brokerage (Rs.)"] = (COL_NET_BROKERAGE, "sum")
        if COL_TURNOVER in df_trades.columns:
            me_agg["Turnover (Rs.)"] = (COL_TURNOVER, "sum")
        monthly_exp = df_trades.groupby(["Month", "Segment"]).agg(**me_agg).reset_index()
        monthly_exp["Month"] = monthly_exp["Month"].astype(str)
        total_cols = [c for c in monthly_exp.columns if "Rs." in c or "Count" in c]
        write_df(ws3, monthly_exp, 4, section="Summary Dashboard", show_total_cols=total_cols)

    # ══════════════════════════════════════════════════════
    # Sheet 4: Turnover Recon
    # ══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Turnover Recon")
    sheet_header(ws4, "Turnover Reconciliation")

    if "Month" in df_trades.columns and COL_TURNOVER in df_trades.columns:
        recon = df_trades.groupby(["Month", "Segment"])[COL_TURNOVER].sum().reset_index()
        recon.columns = ["Month", "Segment", "Turnover - As per LD (Rs.)"]
        recon["Month"] = recon["Month"].astype(str)
        write_df(ws4, recon, 4, section="Turnover Recon", show_total_cols=["Turnover - As per LD (Rs.)"])

    # ══════════════════════════════════════════════════════
    # Sheet 5: Client Concentration
    # ══════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Client Concentration")
    sheet_header(ws5, "Client Concentration Analysis")

    r5 = 4
    if COL_CLIENT_CODE in df_trades.columns:
        nc = COL_CLIENT_NAME if COL_CLIENT_NAME in df_trades.columns else COL_CLIENT_CODE
        cl = df_trades.groupby([COL_CLIENT_CODE, nc]).agg(
            **{"Gross Brokerage (Rs.)": (COL_TOTAL_BROKERAGE, "sum"), "Trade Count": (COL_TOTAL_BROKERAGE, "count")}
        ).reset_index().sort_values("Gross Brokerage (Rs.)", ascending=False)
        cl["% of Total"] = (cl["Gross Brokerage (Rs.)"] / cl["Gross Brokerage (Rs.)"].sum() * 100).round(3)
        cl["Cumulative %"] = cl["% of Total"].cumsum().round(3)

        # HHI block
        hhi_val = compute_hhi(cl["% of Total"])
        kv_block(ws5, r5, [
            ("Total Clients", len(cl)),
            ("HHI Score", round(hhi_val, 0)),
            ("HHI Classification", classify_hhi(hhi_val)),
            (f"Clients > {concentration_pct}%", int((cl["% of Total"] > concentration_pct).sum())),
        ])
        r5 += 7
        sec_title(ws5, r5, f"Top {top_n} Clients by Brokerage")
        r5 += 1
        r5 = write_df(ws5, cl.head(top_n), r5, section="Client Concentration",
                      show_total_cols=["Gross Brokerage (Rs.)", "Trade Count"])

        # Top N by Turnover
        if COL_TURNOVER in df_trades.columns:
            cl_tv = df_trades.groupby([COL_CLIENT_CODE, nc]).agg(**{
                "Turnover (Rs.)": (COL_TURNOVER, "sum"),
                "Gross Brokerage (Rs.)": (COL_TOTAL_BROKERAGE, "sum"),
                "Trade Count": (COL_TOTAL_BROKERAGE, "count"),
            }).reset_index().sort_values("Turnover (Rs.)", ascending=False)
            cl_tv["Avg Rate (bps)"] = np.where(
                cl_tv["Turnover (Rs.)"] > 0,
                cl_tv["Gross Brokerage (Rs.)"] / cl_tv["Turnover (Rs.)"] * 10000, 0
            ).round(2)
            r5 += 1
            sec_title(ws5, r5, f"Top {top_n} Clients by Turnover")
            r5 += 1
            r5 = write_df(ws5, cl_tv.head(top_n), r5,
                          show_total_cols=["Turnover (Rs.)", "Gross Brokerage (Rs.)", "Trade Count"])

        # MoM Brokerage vs Turnover
        if "Month" in df_trades.columns and COL_TURNOVER in df_trades.columns:
            mom_xl = df_trades.groupby("Month").agg(**{
                "Gross Brokerage (Rs.)": (COL_TOTAL_BROKERAGE, "sum"),
                "Turnover (Rs.)": (COL_TURNOVER, "sum"),
            }).reset_index().sort_index()
            mom_xl["Brokerage MoM %"] = mom_variance(mom_xl["Gross Brokerage (Rs.)"]).round(1)
            mom_xl["Turnover MoM %"] = mom_variance(mom_xl["Turnover (Rs.)"]).round(1)
            mom_xl["Avg Rate (bps)"] = np.where(
                mom_xl["Turnover (Rs.)"] > 0,
                mom_xl["Gross Brokerage (Rs.)"] / mom_xl["Turnover (Rs.)"] * 10000, 0
            ).round(2)
            mom_xl["Month"] = mom_xl["Month"].astype(str)
            r5 += 1
            sec_title(ws5, r5, "Month-on-Month: Brokerage vs Turnover")
            r5 += 1
            write_df(ws5, mom_xl, r5,
                     show_total_cols=["Gross Brokerage (Rs.)", "Turnover (Rs.)"])

    # ══════════════════════════════════════════════════════
    # Sheet 6: Scrip Analysis
    # ══════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Scrip Analysis")
    sheet_header(ws6, "Scrip-wise Analysis")

    scn = COL_SCRIP_NAME if COL_SCRIP_NAME in df_trades.columns else COL_SCRIP_CODE
    if scn in df_trades.columns:
        sa = {"Gross Brokerage (Rs.)": (COL_TOTAL_BROKERAGE, "sum"), "Trade Count": (COL_TOTAL_BROKERAGE, "count")}
        if COL_TURNOVER in df_trades.columns:
            sa["Turnover (Rs.)"] = (COL_TURNOVER, "sum")
        se = df_trades.groupby(scn).agg(**sa).reset_index().sort_values("Gross Brokerage (Rs.)", ascending=False)
        if "Turnover (Rs.)" in se.columns:
            se["Avg Rate (bps)"] = np.where(se["Turnover (Rs.)"] > 0,
                                            se["Gross Brokerage (Rs.)"] / se["Turnover (Rs.)"] * 10000, 0).round(2)
        write_df(ws6, se.head(30), 4, section="Scrip Analysis",
                 show_total_cols=["Gross Brokerage (Rs.)", "Trade Count", "Turnover (Rs.)"])

    # ══════════════════════════════════════════════════════
    # Sheet 7: Brokerage Rates
    # ══════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Brokerage Rates")
    sheet_header(ws7, "Brokerage Rate Analytics")

    if "Brokerage_Rate" in df_trades.columns:
        re = df_trades[df_trades["Brokerage_Rate"] > 0].groupby("Segment")["Brokerage_Rate"].agg(
            ["mean", "median", "min", "max", "std", "count"]
        ).reset_index()
        re.columns = ["Segment", "Mean (bps)", "Median (bps)", "Min (bps)", "Max (bps)", "Std Dev (bps)", "Trade Count"]
        for c in ["Mean (bps)", "Median (bps)", "Min (bps)", "Max (bps)", "Std Dev (bps)"]:
            re[c] = (re[c] * 10000).round(4)
        write_df(ws7, re, 4, section="Brokerage Rates")

    # ══════════════════════════════════════════════════════
    # Sheet 8: Error Trades
    # ══════════════════════════════════════════════════════
    ws8 = wb.create_sheet("Error Trades")
    sheet_header(ws8, "Error / Cancelled Trade Analysis")

    r8 = 4
    kv_block(ws8, r8, [
        ("Total Error Trades", total_errors),
        ("Error Rate (%)", round(error_rate, 3)),
        ("Error Brokerage (Rs.)", round(abs(df_errors[COL_TOTAL_BROKERAGE].sum()) if COL_TOTAL_BROKERAGE in df_errors.columns else 0, 2)),
    ])
    r8 += 6
    if not df_errors.empty and "Month" in df_errors.columns:
        sec_title(ws8, r8, "Monthly Error Summary")
        r8 += 1
        ec = COL_TOTAL_BROKERAGE if COL_TOTAL_BROKERAGE in df_errors.columns else df_errors.columns[0]
        ee = df_errors.groupby("Month").agg(**{"Error Count": (ec, "count"), "Error Brokerage (Rs.)": (ec, "sum")}).reset_index()
        ee["Month"] = ee["Month"].astype(str)
        ee["Error Rate (%)"] = (ee["Error Count"] / max(total_trades, 1) * 100).round(4)
        write_df(ws8, ee, r8, section="Error Trades", show_total_cols=["Error Count", "Error Brokerage (Rs.)"])
    else:
        ws8.cell(r8, 1, "No error trade data available.").font = note_font

    # ══════════════════════════════════════════════════════
    # Sheet 9: GST Analytics
    # ══════════════════════════════════════════════════════
    ws9 = wb.create_sheet("GST Analytics")
    sheet_header(ws9, "GST Verification")

    r9 = 4
    gst_base_col_xl = COL_NET_BROKERAGE if COL_NET_BROKERAGE in df_trades.columns else COL_TOTAL_BROKERAGE
    gst_base_lbl_xl = "Net Brokerage" if gst_base_col_xl == COL_NET_BROKERAGE else "Total Brokerage"
    if "Month" in df_trades.columns and "GST_Total" in df_trades.columns:
        total_cgst_x = df_trades[COL_CGST].sum() if COL_CGST in df_trades.columns else 0
        total_sgst_x = df_trades[COL_SGST].sum() if COL_SGST in df_trades.columns else 0
        total_igst_x = df_trades[COL_IGST].sum() if COL_IGST in df_trades.columns else 0
        kv_block(ws9, r9, [
            ("Total CGST (Rs.)", round(total_cgst_x, 2)),
            ("Total SGST (Rs.)", round(total_sgst_x, 2)),
            ("Total IGST (Rs.)", round(total_igst_x, 2)),
            ("Interstate %", round(safe_div(total_igst_x, total_cgst_x + total_sgst_x + total_igst_x) * 100, 1)),
            ("GST Base Column", gst_base_lbl_xl),
        ])
        r9 += 8
        sec_title(ws9, r9, f"Monthly GST Reconciliation (Base: {gst_base_lbl_xl})")
        r9 += 1
        ge = df_trades.groupby("Month").agg(**{
            f"{gst_base_lbl_xl} (Rs.)": (gst_base_col_xl, "sum"),
            "GST Actual (Rs.)": ("GST_Total", "sum"),
        }).reset_index()
        ge["GST Expected @18% (Rs.)"] = ge[f"{gst_base_lbl_xl} (Rs.)"] * GST_RATE
        ge["Difference (Rs.)"] = ge["GST Actual (Rs.)"] - ge["GST Expected @18% (Rs.)"]
        ge["Month"] = ge["Month"].astype(str)
        r9 = write_df(ws9, ge, r9, section="GST Analytics",
                      show_total_cols=[f"{gst_base_lbl_xl} (Rs.)", "GST Actual (Rs.)", "GST Expected @18% (Rs.)", "Difference (Rs.)"])

        # CGST/SGST split verification
        if COL_CGST in df_trades.columns and COL_SGST in df_trades.columns and COL_IGST in df_trades.columns:
            intra_xl = df_trades[
                (df_trades[COL_IGST] == 0) & (df_trades[gst_base_col_xl] > 0) &
                ((df_trades[COL_CGST] > 0) | (df_trades[COL_SGST] > 0))
            ].copy()
            if not intra_xl.empty:
                intra_xl["CGST_Rate"] = (intra_xl[COL_CGST] / intra_xl[gst_base_col_xl] * 100).round(2)
                intra_xl["SGST_Rate"] = (intra_xl[COL_SGST] / intra_xl[gst_base_col_xl] * 100).round(2)
                intra_xl["CGST_OK"] = (intra_xl["CGST_Rate"] - 9.0).abs() <= 0.3
                intra_xl["SGST_OK"] = (intra_xl["SGST_Rate"] - 9.0).abs() <= 0.3
                gst_split_dev_xl = intra_xl[~(intra_xl["CGST_OK"] & intra_xl["SGST_OK"])]
                if not gst_split_dev_xl.empty:
                    r9 += 1
                    sec_title(ws9, r9, "CGST/SGST Split Deviations")
                    r9 += 1
                    gst_split_exp = gst_split_dev_xl[
                        [c for c in [COL_CLIENT_NAME, COL_TXN_DATE, gst_base_col_xl,
                                     COL_CGST, COL_SGST, "CGST_Rate", "SGST_Rate"]
                         if c in gst_split_dev_xl.columns]
                    ].head(50)
                    write_df(ws9, gst_split_exp, r9)

    # ══════════════════════════════════════════════════════
    # Sheet 10: STT Verification
    # ══════════════════════════════════════════════════════
    ws10 = wb.create_sheet("STT Verification")
    sheet_header(ws10, "STT Rate Verification")

    if COL_STT in df_trades.columns:
        sv = COL_MARKET_VALUE if COL_MARKET_VALUE in df_trades.columns else COL_TURNOVER
        se10 = df_trades.groupby("Segment").agg(**{
            "Total STT (Rs.)": (COL_STT, "sum"),
            "Total Value (Rs.)": (sv, "sum"),
            "Trade Count": (COL_STT, "count"),
        }).reset_index()
        se10["Effective Rate (bps)"] = (se10["Total STT (Rs.)"] / se10["Total Value (Rs.)"] * 10000).round(4)
        r10 = write_df(ws10, se10, 4, section="STT Verification",
                       show_total_cols=["Total STT (Rs.)", "Total Value (Rs.)", "Trade Count"])

        # Per-transaction STT deviations for cash delivery
        if COL_MARKET_VALUE in df_trades.columns:
            stt_cash_xl = df_trades[
                (df_trades[COL_STT] > 0) & (df_trades[COL_MARKET_VALUE] > 0)
            ].copy()
            exp_rate = STT_RATES.get("Cash Delivery Buy", 0.001)
            stt_cash_xl["STT_Rate"] = stt_cash_xl[COL_STT] / stt_cash_xl[COL_MARKET_VALUE]
            stt_cash_xl["Rate_OK"] = (stt_cash_xl["STT_Rate"] - exp_rate).abs() <= 0.0002
            stt_dev_xl = stt_cash_xl[~stt_cash_xl["Rate_OK"]]
            if not stt_dev_xl.empty:
                r10 += 1
                sec_title(ws10, r10, "Per-Transaction STT Deviations")
                r10 += 1
                stt_exp_cols = [c for c in [COL_CLIENT_NAME, COL_TXN_DATE, COL_MARKET_VALUE,
                                            COL_STT, "STT_Rate"] if c in stt_dev_xl.columns]
                stt_dev_xl["STT Rate %"] = (stt_dev_xl["STT_Rate"] * 100).round(4)
                write_df(ws10, stt_dev_xl[stt_exp_cols + ["STT Rate %"]].head(50), r10)

    # ══════════════════════════════════════════════════════
    # Sheet 11: SEBI Fees
    # ══════════════════════════════════════════════════════
    ws11 = wb.create_sheet("SEBI Fees")
    sheet_header(ws11, "SEBI Fees Verification")

    sebi_cn = find_column(df_trades, COL_SEBI_FEES)
    if sebi_cn and "Month" in df_trades.columns:
        se11 = df_trades.groupby("Month").agg(**{
            "SEBI Fees (Rs.)": (sebi_cn, "sum"),
            "Turnover (Rs.)": (COL_TURNOVER, "sum"),
        }).reset_index()
        se11["Expected @Rs10/Cr (Rs.)"] = se11["Turnover (Rs.)"] * SEBI_FEE_RATE
        se11["Difference (Rs.)"] = se11["SEBI Fees (Rs.)"] - se11["Expected @Rs10/Cr (Rs.)"]
        se11["Effective Rate (per Cr)"] = (se11["SEBI Fees (Rs.)"] / se11["Turnover (Rs.)"] * 1e7).round(2)
        se11["Month"] = se11["Month"].astype(str)
        r11 = write_df(ws11, se11, 4, section="SEBI Fees",
                       show_total_cols=["SEBI Fees (Rs.)", "Turnover (Rs.)", "Expected @Rs10/Cr (Rs.)", "Difference (Rs.)"])

        # Per-transaction SEBI fee deviations
        if COL_MARKET_VALUE in df_trades.columns:
            sebi_ch_xl = df_trades[
                (df_trades[sebi_cn] > 0) & (df_trades[COL_MARKET_VALUE] > 0)
            ].copy()
            sebi_ch_xl["SEBI_Rate"] = sebi_ch_xl[sebi_cn] / sebi_ch_xl[COL_MARKET_VALUE]
            sebi_ch_xl["Rate_OK"] = (sebi_ch_xl["SEBI_Rate"] - SEBI_FEE_RATE).abs() <= (SEBI_FEE_RATE * 0.1)
            sebi_dv_xl = sebi_ch_xl[~sebi_ch_xl["Rate_OK"]]
            if not sebi_dv_xl.empty:
                r11 += 1
                sec_title(ws11, r11, "Per-Transaction SEBI Fee Deviations")
                r11 += 1
                sebi_dv_cols = [c for c in [COL_CLIENT_NAME, COL_TXN_DATE, COL_MARKET_VALUE,
                                            sebi_cn, "SEBI_Rate"] if c in sebi_dv_xl.columns]
                write_df(ws11, sebi_dv_xl[sebi_dv_cols].head(50), r11)

    # ══════════════════════════════════════════════════════
    # Sheet 12: Stamp Duty
    # ══════════════════════════════════════════════════════
    ws12 = wb.create_sheet("Stamp Duty")
    sheet_header(ws12, "Stamp Duty Analysis")

    stcn = find_column(df_trades, COL_STAMP_CHARGES)
    if stcn and COL_EXCHANGE in df_trades.columns:
        stc = COL_TURNOVER if COL_TURNOVER in df_trades.columns else COL_MARKET_VALUE
        se12 = df_trades.groupby(COL_EXCHANGE).agg(**{
            "Stamp Duty (Rs.)": (stcn, "sum"),
            "Turnover (Rs.)": (stc, "sum"),
        }).reset_index()
        se12["Rate (bps)"] = (se12["Stamp Duty (Rs.)"] / se12["Turnover (Rs.)"] * 10000).round(4)
        write_df(ws12, se12, 4, section="Stamp Duty", show_total_cols=["Stamp Duty (Rs.)", "Turnover (Rs.)"])

    # ══════════════════════════════════════════════════════
    # Sheet 13: Adjustments
    # ══════════════════════════════════════════════════════
    ws13 = wb.create_sheet("Adjustments")
    sheet_header(ws13, "Adjustment Analysis")

    r13 = 4
    if not df_summary.empty:
        adj_ex = df_summary[df_summary["Is_Adj"] == True][["Segment", "Month", "Gross_Brokerage", "Net_Brokerage"]].copy()
        adj_ex.columns = ["Segment", "Month", "Adjustment - Gross (Rs.)", "Adjustment - Net (Rs.)"]
        if not adj_ex.empty:
            non_adj_gross = df_summary[~df_summary["Is_Adj"]]["Gross_Brokerage"].sum()
            adj_total = adj_ex["Adjustment - Gross (Rs.)"].sum()
            kv_block(ws13, r13, [
                ("Total Adjustments (Rs.)", round(adj_total, 2)),
                ("Gross Brokerage (Summary) (Rs.)", round(non_adj_gross, 2)),
                ("Adjustments as % of Gross", round(safe_div(abs(adj_total), non_adj_gross) * 100, 3)),
            ])
            r13 += 6
            sec_title(ws13, r13, "Adjustment Details")
            r13 += 1
            write_df(ws13, adj_ex, r13, section="Adjustments",
                     show_total_cols=["Adjustment - Gross (Rs.)", "Adjustment - Net (Rs.)"])
        else:
            ws13.cell(r13, 1, "No adjustments found in summary sheet.").font = note_font

    # ══════════════════════════════════════════════════════
    # Sheet 14: Buy Sell
    # ══════════════════════════════════════════════════════
    ws14 = wb.create_sheet("Buy Sell")
    sheet_header(ws14, "Buy / Sell Analysis")

    r14 = 4
    if COL_BUY_SELL in df_trades.columns and "Month" in df_trades.columns:
        dft = df_trades.copy()
        dft["Side"] = dft[COL_BUY_SELL].astype(str).str.strip().str.upper()
        bv = COL_MARKET_VALUE if COL_MARKET_VALUE in df_trades.columns else COL_TOTAL_BROKERAGE
        buy_val = dft.loc[dft["Side"].str.startswith("B"), bv].sum()
        sell_val = dft.loc[dft["Side"].str.startswith("S"), bv].sum()
        kv_block(ws14, r14, [
            ("Total Buy Value (Rs.)", round(buy_val, 2)),
            ("Total Sell Value (Rs.)", round(sell_val, 2)),
            ("Buy:Sell Ratio", round(safe_div(buy_val, sell_val), 2)),
        ])
        r14 += 6
        sec_title(ws14, r14, "Monthly Buy / Sell Breakup")
        r14 += 1
        bse = dft.groupby(["Month", "Side"]).agg(**{
            "Trade Count": (COL_TOTAL_BROKERAGE, "count"), "Value (Rs.)": (bv, "sum")
        }).reset_index()
        bse["Month"] = bse["Month"].astype(str)
        write_df(ws14, bse, r14, section="Buy Sell", show_total_cols=["Trade Count", "Value (Rs.)"])

    # ══════════════════════════════════════════════════════
    # Sheet 15: Temporal
    # ══════════════════════════════════════════════════════
    ws15 = wb.create_sheet("Temporal")
    sheet_header(ws15, "Temporal Analysis")

    r15 = 4
    if COL_TXN_DATE in df_trades.columns:
        sec_title(ws15, r15, "Day-of-Week Distribution")
        r15 += 1
        dow_df = df_trades["Day_of_week"].value_counts().reset_index()
        dow_df.columns = ["Day", "Trade Count"]
        dow_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        dow_df["_ord"] = dow_df["Day"].map({d: i for i, d in enumerate(dow_order)})
        dow_df = dow_df.sort_values("_ord").drop(columns="_ord")
        r15 = write_df(ws15, dow_df, r15, section="Temporal", show_total_cols=["Trade Count"])

        # Exchange holiday trades
        hol_dates_xl = pd.to_datetime(FY2026_HOLIDAYS)
        hol_trades_xl = df_trades[df_trades[COL_TXN_DATE].dt.normalize().isin(hol_dates_xl)]
        if not hol_trades_xl.empty:
            r15 += 1
            sec_title(ws15, r15, "Trades on Exchange Holidays")
            r15 += 1
            hol_sum_xl = hol_trades_xl.groupby(COL_TXN_DATE).agg(
                **{"Trade Count": (COL_TOTAL_BROKERAGE, "count"),
                   "Value (Rs.)": (COL_MARKET_VALUE, "sum") if COL_MARKET_VALUE in hol_trades_xl.columns else (COL_TOTAL_BROKERAGE, "sum")},
            ).reset_index()
            hol_sum_xl["Day"] = hol_sum_xl[COL_TXN_DATE].dt.day_name()
            r15 = write_df(ws15, hol_sum_xl, r15, show_total_cols=["Trade Count", "Value (Rs.)"])

        # Daily turnover ranking
        if COL_TURNOVER in df_trades.columns:
            r15 += 1
            sec_title(ws15, r15, "Daily Turnover Ranking (Top Days for OTR Sampling)")
            r15 += 1
            dt_xl = df_trades.groupby(COL_TXN_DATE).agg(**{
                "Turnover (Rs.)": (COL_TURNOVER, "sum"),
                "Trade Count": (COL_TOTAL_BROKERAGE, "count"),
                "Brokerage (Rs.)": (COL_TOTAL_BROKERAGE, "sum"),
            }).reset_index().sort_values("Turnover (Rs.)", ascending=False)
            dt_xl["Rank"] = range(1, len(dt_xl) + 1)
            dt_xl["Day"] = dt_xl[COL_TXN_DATE].dt.day_name()
            write_df(ws15, dt_xl.head(top_n)[["Rank", COL_TXN_DATE, "Day", "Trade Count",
                                               "Turnover (Rs.)", "Brokerage (Rs.)"]],
                     r15, show_total_cols=["Turnover (Rs.)", "Brokerage (Rs.)"])

    # ══════════════════════════════════════════════════════
    # Sheet 16: Client × Product Analysis
    # ══════════════════════════════════════════════════════
    ws_cp = wb.create_sheet("Client Product")
    sheet_header(ws_cp, "Client × Product Analysis")

    rcp = 4
    if COL_CLIENT_CODE in df_trades.columns and "Segment" in df_trades.columns:
        nc_xl = COL_CLIENT_NAME if COL_CLIENT_NAME in df_trades.columns else COL_CLIENT_CODE
        cp_agg_xl = {"Gross Brokerage (Rs.)": (COL_TOTAL_BROKERAGE, "sum"),
                     "Trade Count": (COL_TOTAL_BROKERAGE, "count")}
        if COL_TURNOVER in df_trades.columns:
            cp_agg_xl["Turnover (Rs.)"] = (COL_TURNOVER, "sum")
        cp_xl = df_trades.groupby([COL_CLIENT_CODE, nc_xl, "Segment"]).agg(**cp_agg_xl).reset_index()
        cp_xl = cp_xl.sort_values("Gross Brokerage (Rs.)", ascending=False)
        if "Turnover (Rs.)" in cp_xl.columns:
            cp_xl["Avg Rate (bps)"] = np.where(
                cp_xl["Turnover (Rs.)"] > 0,
                cp_xl["Gross Brokerage (Rs.)"] / cp_xl["Turnover (Rs.)"] * 10000, 0
            ).round(2)

        # Revenue mix summary
        seg_rev_xl = cp_xl.groupby("Segment").agg(**{
            "Gross Brokerage (Rs.)": ("Gross Brokerage (Rs.)", "sum"),
            "Clients": (COL_CLIENT_CODE, "nunique"),
            "Trades": ("Trade Count", "sum"),
        }).reset_index()
        seg_rev_xl["% of Revenue"] = (
            seg_rev_xl["Gross Brokerage (Rs.)"] / seg_rev_xl["Gross Brokerage (Rs.)"].sum() * 100
        ).round(2)
        sec_title(ws_cp, rcp, "Revenue Mix by Product Segment")
        rcp += 1
        rcp = write_df(ws_cp, seg_rev_xl, rcp,
                        show_total_cols=["Gross Brokerage (Rs.)", "Trades"])
        rcp += 1

        # Top clients by segment
        sec_title(ws_cp, rcp, f"Top {top_n} Clients — Product-wise Detail")
        rcp += 1
        # Get top N clients
        ct_xl = cp_xl.groupby(COL_CLIENT_CODE)["Gross Brokerage (Rs.)"].sum().nlargest(top_n).index
        cp_top_xl = cp_xl[cp_xl[COL_CLIENT_CODE].isin(ct_xl)]
        rcp = write_df(ws_cp, cp_top_xl, rcp, section="Client Product",
                      show_total_cols=["Gross Brokerage (Rs.)", "Trade Count",
                                       "Turnover (Rs.)"] if "Turnover (Rs.)" in cp_xl.columns
                      else ["Gross Brokerage (Rs.)", "Trade Count"])

        # Product Type Rate Uniformity (non-uniform rates)
        if COL_PRODUCT_DESC in df_trades.columns and COL_MARKET_VALUE in df_trades.columns:
            pru_xl = df_trades[
                (df_trades[COL_TOTAL_BROKERAGE] > 0) & (df_trades[COL_MARKET_VALUE] > 0)
            ].copy()
            pru_xl["Rate"] = pru_xl[COL_TOTAL_BROKERAGE] / pru_xl[COL_MARKET_VALUE]
            nc_xl2 = COL_CLIENT_NAME if COL_CLIENT_NAME in df_trades.columns else COL_CLIENT_CODE
            ru_xl = pru_xl.groupby([COL_CLIENT_CODE, nc_xl2, COL_PRODUCT_DESC]).agg(**{
                "Trade Count": ("Rate", "count"),
                "Min Rate (bps)": ("Rate", "min"),
                "Max Rate (bps)": ("Rate", "max"),
                "Gross Brokerage (Rs.)": (COL_TOTAL_BROKERAGE, "sum"),
            }).reset_index()
            ru_xl["Min Rate (bps)"] = (ru_xl["Min Rate (bps)"] * 10000).round(2)
            ru_xl["Max Rate (bps)"] = (ru_xl["Max Rate (bps)"] * 10000).round(2)
            ru_xl["Range (bps)"] = (ru_xl["Max Rate (bps)"] - ru_xl["Min Rate (bps)"]).round(2)
            nu_xl = ru_xl[(ru_xl["Range (bps)"] > 1) & (ru_xl["Trade Count"] > 1)].sort_values(
                "Range (bps)", ascending=False
            )
            if not nu_xl.empty:
                rcp += 1
                sec_title(ws_cp, rcp, "Non-Uniform Brokerage Rates by Client × Product")
                rcp += 1
                rcp = write_df(ws_cp, nu_xl.head(50), rcp,
                               show_total_cols=["Gross Brokerage (Rs.)"])

        # Product Type Distribution
        if COL_PRODUCT_DESC in df_trades.columns:
            pd_xl = df_trades.groupby(COL_PRODUCT_DESC).agg(**{
                "Trade Count": (COL_TOTAL_BROKERAGE, "count"),
                "Gross Brokerage (Rs.)": (COL_TOTAL_BROKERAGE, "sum"),
            }).reset_index().sort_values("Gross Brokerage (Rs.)", ascending=False)
            pd_xl["% of Brokerage"] = (
                pd_xl["Gross Brokerage (Rs.)"] / pd_xl["Gross Brokerage (Rs.)"].sum() * 100
            ).round(2)
            rcp += 1
            sec_title(ws_cp, rcp, "Product Type Distribution")
            rcp += 1
            write_df(ws_cp, pd_xl, rcp,
                     show_total_cols=["Gross Brokerage (Rs.)", "Trade Count"])

    # ══════════════════════════════════════════════════════
    # Sheet 17: Transaction-Level Analysis
    # ══════════════════════════════════════════════════════
    ws_txn = wb.create_sheet("Txn Analysis")
    sheet_header(ws_txn, "Transaction-Level Analysis")

    rtx = 4
    if not df_trades.empty and COL_TOTAL_BROKERAGE in df_trades.columns:
        # Distribution stats
        brok_s = df_trades[COL_TOTAL_BROKERAGE]
        kv_block(ws_txn, rtx, [
            ("Total Trades", total_trades),
            ("Mean Brokerage (Rs.)", round(brok_s.mean(), 2)),
            ("Median Brokerage (Rs.)", round(brok_s.median(), 2)),
            ("Std Dev (Rs.)", round(brok_s.std(), 2)),
            ("Max Brokerage (Rs.)", round(brok_s.max(), 2)),
            ("Min Brokerage (Rs.)", round(brok_s.min(), 2)),
            ("Skewness", round(brok_s.skew(), 2)),
            ("Zero Brokerage Trades", int((brok_s == 0).sum())),
            ("Negative Brokerage Trades", int((brok_s < 0).sum())),
            ("Trades >= Materiality", int((brok_s.abs() >= materiality).sum())),
        ])
        rtx += 14

        # Top trades
        sec_title(ws_txn, rtx, f"Top {top_n} Trades by Brokerage")
        rtx += 1
        top_cols_xl = [c for c in [COL_TXN_DATE, COL_CLIENT_CODE, COL_CLIENT_NAME,
                                    COL_SCRIP_NAME, "Segment", COL_BUY_SELL,
                                    COL_TURNOVER, COL_TOTAL_BROKERAGE]
                       if c in df_trades.columns]
        top_xl = df_trades.nlargest(top_n, COL_TOTAL_BROKERAGE)[top_cols_xl]
        rtx = write_df(ws_txn, top_xl, rtx, section="Transaction Analysis",
                        show_total_cols=[COL_TOTAL_BROKERAGE, COL_TURNOVER])

        # SEBI Max 2.5% violations
        if "Brokerage_Rate" in df_trades.columns:
            over_xl = df_trades[df_trades["Brokerage_Rate"] > SEBI_MAX_BROKERAGE_PCT]
            if not over_xl.empty:
                rtx += 1
                sec_title(ws_txn, rtx, "Trades Exceeding SEBI Max 2.5% Brokerage")
                rtx += 1
                over_exp = over_xl[top_cols_xl].head(50).copy()
                over_exp["Rate %"] = (over_xl["Brokerage_Rate"].head(50) * 100).round(4).values
                rtx = write_df(ws_txn, over_exp, rtx)

        # High Turnover / Low Brokerage
        if COL_TURNOVER in df_trades.columns and "Brokerage_Rate" in df_trades.columns:
            vht = df_trades[(df_trades[COL_TURNOVER] > 0) & (df_trades["Brokerage_Rate"] > 0)]
            if len(vht) > 20:
                tv75 = vht[COL_TURNOVER].quantile(0.75)
                rt10 = vht["Brokerage_Rate"].quantile(0.10)
                htlb_xl = vht[(vht[COL_TURNOVER] >= tv75) & (vht["Brokerage_Rate"] <= rt10)]
                if not htlb_xl.empty:
                    rtx += 1
                    sec_title(ws_txn, rtx, "High Turnover / Low Brokerage Trades")
                    rtx += 1
                    htlb_exp = htlb_xl[top_cols_xl].head(50).copy()
                    htlb_exp["Rate (bps)"] = (htlb_xl["Brokerage_Rate"].head(50) * 10000).round(2).values
                    rtx = write_df(ws_txn, htlb_exp, rtx)

        # Benford's Law
        brok_pos_xl = df_trades[COL_TOTAL_BROKERAGE][df_trades[COL_TOTAL_BROKERAGE] > 0]
        if len(brok_pos_xl) > 100:
            rtx += 1
            sec_title(ws_txn, rtx, "Benford's Law — First Digit Analysis")
            rtx += 1
            fd_xl = brok_pos_xl.apply(
                lambda x: int(str(x).lstrip("0").replace(".", "")[0]) if x > 0 else 0
            )
            fd_xl = fd_xl[fd_xl > 0]
            obs_xl = fd_xl.value_counts(normalize=True).sort_index()
            bf_df_xl = pd.DataFrame({
                "Digit": range(1, 10),
                "Expected %": [round(math.log10(1 + 1 / d) * 100, 2) for d in range(1, 10)],
                "Observed %": [round(obs_xl.get(d, 0) * 100, 2) for d in range(1, 10)],
            })
            bf_df_xl["Deviation %"] = (bf_df_xl["Observed %"] - bf_df_xl["Expected %"]).round(2)
            bf_df_xl["MAD"] = bf_df_xl["Deviation %"].abs().mean()
            write_df(ws_txn, bf_df_xl, rtx)

    # ══════════════════════════════════════════════════════
    # Sheet 18: All Audit Procedures (Master)
    # ══════════════════════════════════════════════════════
    ws16 = wb.create_sheet("All Audit Procedures")
    sheet_header(ws16, "Master Audit Procedures")

    hdr_row(ws16, 4, ["Sr. No.", "Area", "Procedure", "Expected Evidence", "Auditor Remarks"])
    r = 5
    sr = 1
    for sname, procs in AUDIT_PROCEDURES.items():
        # Section separator row
        ws16.cell(r, 1).fill = proc_hdr_fill
        ws16.cell(r, 2, sname).font = body_b
        ws16.cell(r, 2).fill = proc_hdr_fill
        for cc in range(3, 6):
            ws16.cell(r, cc).fill = proc_hdr_fill
        r += 1
        for sno, proc, ev in procs:
            data_row(ws16, r, [sr, sname, proc, ev, ""], stripe=(sr % 2 == 0))
            ws16.cell(r, 3).alignment = wrap
            ws16.cell(r, 4).alignment = wrap
            ws16.cell(r, 5).alignment = wrap
            r += 1
            sr += 1

    ws16.column_dimensions["A"].width = 8
    ws16.column_dimensions["B"].width = 22
    ws16.column_dimensions["C"].width = 65
    ws16.column_dimensions["D"].width = 42
    ws16.column_dimensions["E"].width = 30

    # ── Cleanup ──
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# Generate and offer download
if st.button("Generate Excel Report", type="primary"):
    with st.spinner("Generating report..."):
        excel_buffer = generate_excel_report()
    filename = f"Brokerage_Analytics_{entity_name.replace(' ', '_')}_{audit_period.replace(' ', '_')}.xlsx"
    st.download_button(
        label="Download Excel Report",
        data=excel_buffer,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.success("Report generated successfully!")


# ═══════════════════════════════════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════════════════════════════════
st.markdown(
    '<div class="footer-text">'
    f'{APP_FULL_NAME} — KKC & Associates LLP, Chartered Accountants<br>'
    f'Generated: {datetime.now().strftime("%d %B %Y")}'
    '</div>',
    unsafe_allow_html=True,
)
